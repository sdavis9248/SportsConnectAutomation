"""
eTrainu  <->  Compliance matcher
================================

Given (a) the resolved volunteer+certification rows from the compliance provider
and (b) the available eTrainu training events, produce a per-volunteer worklist:

    volunteer  ->  deficient cert  ->  next step to close the gap

The "next step" is one of three channels:

  * etrainu : an in-person AYSO course (coach / referee / AD training). We find the
              SOONEST qualifying session and hand back its date / time / location /
              enroll handle / contact.
  * portal  : an ONLINE cert (SafeSport, Safe Haven, Concussion, Cardiac, LiveScan
              fingerprinting, JDP background check). These are NOT eTrainu events;
              they live in separate portals, so we return a static link instead.
  * admin   : something only a region admin can clear in the Governing System
              (e.g. ID verification). The volunteer cannot self-serve.

Design goals (consistent with the rest of the repo):
  * Pure engine (`build_remediation`) decoupled from any output sink, so it is
    trivially testable and can feed either a standalone worklist OR the board
    portal's compliance tab.
  * Loaders that accept BOTH the compliance_test harness output
    (compliance_resolved.json) and the portal payload (compliance.json), and
    BOTH the live-scrape event JSON and a hand-built events list.
  * Every mapping table is overridable from config under
    config['etrainu_compliance'][...] so policy lives in config, not code.

CLI entry is wired in main.py as `--etrainu-compliance` (worklist + portal
next-steps both default-ON, with --no-worklist / --no-portal-next-steps opt-outs).
"""
from __future__ import annotations

import csv
import json
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# Channel classification (which sink closes a deficient cert).
#   portal : self-serve ONLINE training (separate portals, static links)
#   etrainu: in-person AYSO course on the eTrainu calendar
#   admin  : Risk Status / background check, cleared by the region — not self-serve
ONLINE_PORTAL_CERTS = ('safe_haven', 'safesport', 'concussion', 'cardiac', 'fingerprinting')
ETRAINU_CERTS = ('coach_license', 'referee_certification')
RISK_KEY = 'risk_status'   # special: read from the volunteer's top-level risk_status

# ── Region 58 compliance policy (mirrors the Apps Script) ──────────────────────
# Compliance is tracked for FOUR roles only: Head Coach, Assistant Coach, Referee,
# Youth Referee. Everyone else (Manager, Team Parent, board roles, etc.) is NOT
# tracked. Youth Referees are EXEMPT from SafeSport and Fingerprinting.
# Override via config['etrainu_compliance']['required_by_role'].
_CORE = ('safe_haven', 'concussion', 'cardiac', RISK_KEY)   # all four tracked roles
DEFAULT_REQUIRED_BY_ROLE: Dict[str, Tuple[str, ...]] = {
    'coach':          _CORE + ('fingerprinting', 'safesport', 'coach_license'),
    'referee':        _CORE + ('fingerprinting', 'safesport', 'referee_certification'),
    'youth_referee':  _CORE + ('referee_certification',),   # no SafeSport, no Fingerprinting
    'untracked':      (),                                    # not a compliance-tracked role
}

# Position-string keyword -> role bucket. First hit wins, so 'youth referee' MUST
# precede 'referee', and the coach checks precede the bare referee check.
_ROLE_KEYWORDS = (
    ('youth referee', 'youth_referee'),
    ('head coach', 'coach'), ('assistant coach', 'coach'), ('coach', 'coach'),
    ('referee', 'referee'), ('assistant referee', 'referee'), (' ar', 'referee'),
)
TRACKED_BUCKETS = ('coach', 'referee', 'youth_referee')

# ── eTrainu course mapping ─────────────────────────────────────────────────────
# A deficient coach_license is satisfied by the coach course for the division the
# volunteer coaches. Division label (normalized) -> eTrainu course_type string.
# Course_type strings are exactly what etrainu_manager classifies events into.
DEFAULT_COACH_DIVISION_COURSE: Dict[str, str] = {
    '05U': '6U/8U Coach', '5U': '6U/8U Coach',
    '06U': '6U/8U Coach', '6U': '6U/8U Coach',
    '07U': '6U/8U Coach', '7U': '6U/8U Coach',
    '08U': '6U/8U Coach', '8U': '6U/8U Coach',
    '10U': '10U Coach',
    '12U': '12U Coach',
    '14U': '14U/Intermediate Coach',
    '16U': '14U/Intermediate Coach',
    '19U': '14U/Intermediate Coach',
}
# Ordered fallbacks when the division-specific course isn't on the calendar.
_COACH_COURSE_FALLBACKS = ('Coach Certification', '6U/8U Coach', '10U Coach',
                           '12U Coach', '14U/Intermediate Coach')
# A deficient referee_certification: entry-level Regional Referee first.
_REFEREE_COURSE_PREFERENCE = ('Regional Referee', 'Referee Certification',
                              'Intermediate Referee')

# ── Static portal link table for the NON-eTrainu certs ─────────────────────────
# URLs taken from the Region 58 compliance Apps Script (the authoritative source).
# Override any via config['etrainu_compliance']['portal_links'][cert_key] = {...}.
DEFAULT_PORTAL_LINKS: Dict[str, Dict[str, str]] = {
    'safesport':       {'label': 'SafeSport Training',
                        'url': 'https://safesporttrained.org/?KeyName=tsVWe36Xa6PS3b5NzOug',
                        'note': 'Required SafeSport abuse-prevention training.'},
    'safe_haven':      {'label': 'AYSO Safe Haven',
                        'url': 'https://www.aysou.org/',
                        'note': 'Required AYSO Safe Haven certification (AYSOU).'},
    'concussion':      {'label': 'CDC Concussion Awareness',
                        'url': 'https://www.cdc.gov/headsup/youthsports/training/index.html',
                        'note': 'Required concussion awareness certification.'},
    'cardiac':         {'label': 'Sudden Cardiac Arrest',
                        'url': 'https://www.nfhslearn.com/courses/sudden-cardiac-arrest',
                        'note': 'Required cardiac arrest awareness training.'},
    'fingerprinting':  {'label': 'CA Mandated Fingerprinting (LiveScan)',
                        'url': 'https://www.ayso.org/volunteer/livescan/',
                        'note': 'Required background-check fingerprinting (LiveScan).'},
}
# AYSOU is the fallback link for coach/referee certs when no eTrainu session is
# yet on the calendar.
AYSOU_URL = 'https://www.aysou.org/'
# Risk Status is cleared by the region, not self-served.
RISK_ADMIN_NOTE = ('Risk Status / background check is cleared by the region in the '
                   'Governing System after LiveScan + JDP; not a self-serve step.')


# ──────────────────────────────────────────────────────────────────────────────
#  Helpers: normalization
# ──────────────────────────────────────────────────────────────────────────────
def _role_bucket(position: str) -> str:
    p = (position or '').lower()
    for kw, bucket in _ROLE_KEYWORDS:
        if kw in p:
            return bucket
    return 'untracked'


def _extract_age(text: str) -> int:
    """Pull the age number from a division ('10U Boys') or license ('10U Coach').
    Mirrors the Apps Script's extractAgeFromDivision/extractAgeFromLicense."""
    import re
    if not text:
        return 0
    m = re.search(r'(\d+)\s*[uU]', str(text))
    return int(m.group(1)) if m else 0


def _risk_compliance(risk_status: str, bucket: str) -> Tuple[bool, str]:
    """Port of the Apps Script checkRiskStatusCompliance, keyed on the normalized
    role bucket. Returns (compliant, issue_text)."""
    if not risk_status:
        return False, 'Risk Status - background check not completed'
    s = str(risk_status).strip().lower()
    is_coach = bucket == 'coach'
    is_youth = bucket == 'youth_referee'
    if s == 'green':
        return True, ''
    if s == 'blue':
        return (True, '') if is_youth else (False, 'Risk Status - Blue valid only for Youth Referees')
    if s == 'yellow':
        return (False, 'Risk Status - Yellow restricts coaching') if is_coach else (True, '')
    if s == 'orange':
        return True, ''
    if s == 'brown':
        return (False, 'Risk Status - Brown restricts coaching') if is_coach else (True, '')
    if s in ('red', 'disqualified', 'ineligible'):
        return False, 'Risk Status - not eligible to volunteer'
    if s in ('gray', 'grey', 'pending'):
        return False, 'Risk Status - background check pending'
    if 'expired' in s:
        return False, 'Risk Status - background check expired'
    if s in ('none', 'no status', ''):
        return False, 'Risk Status - background check not completed'
    return False, f'Risk Status - unknown status: {risk_status}'



def _norm_division(div: str) -> str:
    """Pull a division token like '10U' out of a free-form division/program name."""
    import re
    if not div:
        return ''
    m = re.search(r'(\d{1,2})\s*[uU]', div)
    if m:
        return f"{int(m.group(1)):02d}U"
    return div.strip().upper()


def _cert_is_deficient(cert: Optional[Dict[str, Any]]) -> Optional[str]:
    """Return a short status string if the cert is a gap, else None (it's fine).
    Missing cert -> 'missing'. Present but bad -> its status. Compliant -> None."""
    if not cert:
        return 'missing'
    status = (cert.get('status') or 'unknown').lower()
    verified = bool(cert.get('verified'))
    if verified and status == 'valid':
        return None
    if status in ('expired', 'invalid', 'pending'):
        return status
    return 'unverified' if status == 'valid' else (status or 'missing')


def required_certs_for(position: str, policy: Dict[str, Tuple[str, ...]]) -> Tuple[str, ...]:
    return policy.get(_role_bucket(position), ())


# ──────────────────────────────────────────────────────────────────────────────
#  Helpers: eTrainu session selection
# ──────────────────────────────────────────────────────────────────────────────
def _event_sessions(event: Dict[str, Any]) -> List[Dict[str, Any]]:
    return event.get('sessions') or []


def _session_date(session: Dict[str, Any]) -> Optional[date]:
    d = session.get('date')
    if not d:
        return None
    try:
        return datetime.strptime(d[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


def next_session_for_course_types(events: List[Dict[str, Any]],
                                  course_types: List[str],
                                  today: date) -> Optional[Dict[str, Any]]:
    """Soonest upcoming session whose event.course_type is in `course_types`
    (tried in priority order). Returns a flat dict ready for the worklist."""
    for ct in course_types:
        best = None  # (date, event, session)
        for ev in events:
            if ev.get('course_type') != ct:
                continue
            for s in _event_sessions(ev):
                sd = _session_date(s)
                if sd is None or sd < today:
                    continue
                if best is None or sd < best[0]:
                    best = (sd, ev, s)
        if best:
            _, ev, s = best
            return {
                'course_type': ct,
                'title': ev.get('title'),
                'date': s.get('date'),
                'day_of_week': s.get('day_of_week'),
                'start_time': s.get('start_time'),
                'end_time': s.get('end_time'),
                'location': s.get('location'),
                'region': ev.get('region'),
                'enroll': ev.get('enroll_info') or {},
                'contact': ev.get('contact') or {},
                'source_url': ev.get('source_url'),
            }
    return None


def _coach_course_priority(division: str,
                           div_course_map: Dict[str, str]) -> List[str]:
    token = _norm_division(division)
    primary = div_course_map.get(token)
    order: List[str] = []
    if primary:
        order.append(primary)
    for fb in _COACH_COURSE_FALLBACKS:
        if fb not in order:
            order.append(fb)
    return order


# ──────────────────────────────────────────────────────────────────────────────
#  Core engine
# ──────────────────────────────────────────────────────────────────────────────
def build_remediation(resolved_volunteers: List[Dict[str, Any]],
                      events: List[Dict[str, Any]],
                      today: Optional[date] = None,
                      config: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """For each volunteer, compute the list of deficient required certs and the
    next step to close each. Returns one remediation dict per volunteer:

        {
          email, first_name, last_name, division, position, role,
          match_confidence, matched (bool),
          gaps: [ {cert, status, channel, course_type?, next_session?, portal?, note?}, ... ],
          summary: "<one-line next action>",
        }

    A volunteer with no gaps still appears (gaps == []), so the worklist can show
    "compliant" rows or callers can filter them out.
    """
    today = today or date.today()
    cfg = (config or {}).get('etrainu_compliance', {}) if config else {}
    policy = {**DEFAULT_REQUIRED_BY_ROLE, **(cfg.get('required_by_role') or {})}
    div_course_map = {**DEFAULT_COACH_DIVISION_COURSE, **(cfg.get('coach_division_course') or {})}
    portal_links = {**DEFAULT_PORTAL_LINKS, **(cfg.get('portal_links') or {})}

    out: List[Dict[str, Any]] = []
    for v in resolved_volunteers:
        email = (v.get('volunteer_email') or v.get('email') or '').strip()
        first = v.get('volunteer_first_name') or v.get('first_name') or ''
        last = v.get('volunteer_last_name') or v.get('last_name') or ''
        position = v.get('volunteer_position') or v.get('position') or ''
        division = (v.get('package_name') or v.get('division')
                    or v.get('Division Name') or '')
        confidence = v.get('_match_confidence') or v.get('match_confidence') or 'none'
        matched = (v.get('_match_method') or v.get('match_method') or 'none') != 'none'
        certs = v.get('certifications') or {}
        risk = v.get('risk_status') or v.get('Risk Status') or ''
        bucket = _role_bucket(position)
        tracked = bucket in TRACKED_BUCKETS

        # Roles outside the four tracked buckets are not assessed at all.
        if not tracked:
            out.append({
                'email': email, 'first_name': first, 'last_name': last,
                'division': division, 'position': position, 'role': bucket,
                'match_confidence': confidence, 'matched': matched, 'tracked': False,
                'gaps': [], 'summary': 'Not a compliance-tracked role.',
            })
            continue

        # Tracked role but no governing-system record: we can't see their certs,
        # so flag for identity verification rather than inventing every gap.
        if not matched:
            out.append({
                'email': email, 'first_name': first, 'last_name': last,
                'division': division, 'position': position, 'role': bucket,
                'match_confidence': confidence, 'matched': False, 'tracked': True,
                'gaps': [], 'summary': 'No governing-system match — verify identity '
                                       'before assessing certifications.',
            })
            continue

        required = required_certs_for(position, policy)
        gaps: List[Dict[str, Any]] = []
        for key in required:
            # Risk Status is a top-level field with role-dependent rules, not a cert.
            if key == RISK_KEY:
                ok, issue = _risk_compliance(risk, bucket)
                if ok:
                    continue
                gaps.append({'cert': 'risk_status', 'status': (risk or 'none'),
                             'channel': 'admin', 'note': issue or RISK_ADMIN_NOTE})
                continue

            status = _cert_is_deficient(certs.get(key))
            # coach_license can be PRESENT but too low for the division — treat an
            # insufficient license as a gap even though the cert itself is "valid".
            insufficient = False
            if key == 'coach_license' and status is None:
                lic = certs.get('coach_license') or {}
                lic_age = _extract_age(lic.get('detail') or '')
                div_age = _extract_age(division)
                if lic_age and div_age and lic_age < div_age:
                    insufficient, status = True, f'insufficient ({lic.get("detail")})'
            if status is None:
                continue  # compliant on this cert

            gap: Dict[str, Any] = {'cert': key, 'status': status}
            if key == 'coach_license':
                gap['channel'] = 'etrainu'
                order = _coach_course_priority(division, div_course_map)
                gap['next_session'] = next_session_for_course_types(events, order, today)
                gap['target_courses'] = order
                if insufficient:
                    gap['note'] = f'License {certs["coach_license"].get("detail")} below {division} requirement'
                gap['fallback_url'] = AYSOU_URL
            elif key == 'referee_certification':
                gap['channel'] = 'etrainu'
                order = list(_REFEREE_COURSE_PREFERENCE)
                gap['next_session'] = next_session_for_course_types(events, order, today)
                gap['target_courses'] = order
                gap['fallback_url'] = AYSOU_URL
            else:
                gap['channel'] = 'portal'
                gap['portal'] = portal_links.get(key, {'label': key, 'url': '', 'note': ''})
            gaps.append(gap)

        out.append({
            'email': email, 'first_name': first, 'last_name': last,
            'division': division, 'position': position, 'role': bucket,
            'match_confidence': confidence, 'matched': True, 'tracked': True,
            'gaps': gaps, 'summary': _summarize(gaps, True, tracked=True),
        })
    n_track = sum(1 for r in out if r.get('tracked'))
    n_gap = sum(1 for r in out if r['gaps'])
    logger.info(f"Remediation: {len(out)} volunteers, {n_track} tracked, {n_gap} with >=1 gap.")
    return out


def _summarize(gaps: List[Dict[str, Any]], matched: bool, tracked: bool = True) -> str:
    if not tracked:
        return 'Not a compliance-tracked role.'
    if not matched:
        return 'No governing-system match — verify identity before assessing certifications.'
    if not gaps:
        return 'Compliant.'
    parts = []
    for g in gaps:
        if g['channel'] == 'etrainu':
            ns = g.get('next_session')
            if ns:
                parts.append(f"{g['cert']}: {ns['course_type']} on {ns.get('date')} "
                             f"{ns.get('start_time') or ''} @ {ns.get('location') or 'TBD'}")
            else:
                parts.append(f"{g['cert']}: no eTrainu session scheduled yet")
        elif g['channel'] == 'portal':
            parts.append(f"{g['cert']}: {g['portal'].get('label', g['cert'])} (online)")
        else:
            parts.append(f"{g['cert']}: region admin action")
    return ' | '.join(parts)


# ──────────────────────────────────────────────────────────────────────────────
#  Loaders (offline-friendly)
# ──────────────────────────────────────────────────────────────────────────────
def load_events(path: str) -> List[Dict[str, Any]]:
    """Load events from a saved scrape. Accepts either a bare list or the
    manager's {'events': [...]} / {'scraped_events': [...]} wrappers. If `path`
    is a directory, picks the newest etrainu_live_events_*.json in it."""
    p = Path(path)
    if p.is_dir():
        hits = sorted(p.glob('etrainu_live_events_*.json'),
                      key=lambda x: x.stat().st_mtime, reverse=True)
        if not hits:
            raise FileNotFoundError(f"No etrainu_live_events_*.json in {p}")
        p = hits[0]
    data = json.loads(p.read_text(encoding='utf-8'))
    if isinstance(data, dict):
        data = data.get('events') or data.get('scraped_events') or []
    logger.info(f"Loaded {len(data)} events from {p}")
    return data


def load_resolved(path: str) -> List[Dict[str, Any]]:
    """Load resolved volunteers from EITHER compliance_resolved.json (a list of
    resolved rows) OR the portal compliance.json ({'volunteers': [...]})."""
    data = json.loads(Path(path).read_text(encoding='utf-8'))
    if isinstance(data, dict) and 'volunteers' in data:
        return data['volunteers']
    if isinstance(data, list):
        return data
    raise ValueError(f"Unrecognized resolved-compliance shape in {path}")


# ──────────────────────────────────────────────────────────────────────────────
#  Output sink #1 — standalone worklist (xlsx + csv)
# ──────────────────────────────────────────────────────────────────────────────
_WORKLIST_HEADERS = [
    'Last Name', 'First Name', 'Email', 'Division', 'Position', 'Role',
    'Match Confidence', 'Deficient Cert', 'Status', 'Channel',
    'Next Step', 'Date', 'Day', 'Time', 'Location', 'Enroll / Link', 'Contact',
]


def _worklist_rows(remediations: List[Dict[str, Any]],
                   include_compliant: bool = False) -> List[List[str]]:
    rows: List[List[str]] = []
    for r in remediations:
        if not r.get('tracked', True):
            continue  # untracked roles never appear on the chase list
        base = [r['last_name'], r['first_name'], r['email'], r['division'],
                r['position'], r['role'], r['match_confidence']]
        if not r['gaps']:
            if not r['matched']:
                # Tracked but unmatched: actionable (verify identity first).
                rows.append(base + ['(identity)', 'unmatched', 'admin', r['summary'],
                                    '', '', '', '', '', ''])
            elif include_compliant:
                rows.append(base + ['(compliant)', '', '', r['summary'],
                                    '', '', '', '', '', ''])
            continue
        for g in r['gaps']:
            channel = g['channel']
            nxt = date_s = day_s = time_s = loc_s = link_s = contact_s = ''
            if channel == 'etrainu':
                ns = g.get('next_session')
                if ns:
                    nxt = ns.get('course_type') or 'eTrainu course'
                    date_s = ns.get('date') or ''
                    day_s = ns.get('day_of_week') or ''
                    time_s = f"{ns.get('start_time','')}-{ns.get('end_time','')}".strip('-')
                    loc_s = ns.get('location') or ''
                    enroll = ns.get('enroll') or {}
                    link_s = (enroll.get('data_event') and
                              f"event={enroll.get('data_event')};session={enroll.get('data_session')}") \
                             or ns.get('source_url') or ''
                    c = ns.get('contact') or {}
                    contact_s = ' '.join(filter(None, [c.get('name'), c.get('email'),
                                                       c.get('phone')]))
                else:
                    targets = '/'.join(g.get('target_courses', [])[:2])
                    nxt = f"Await eTrainu session ({targets})"
                    link_s = g.get('fallback_url') or ''
                if g.get('note'):
                    contact_s = (contact_s + ' | ' + g['note']).strip(' |')
            elif channel == 'portal':
                pl = g.get('portal', {})
                nxt = pl.get('label') or g['cert']
                link_s = pl.get('url') or ''
                contact_s = pl.get('note') or ''
            else:  # admin (risk status / background check)
                nxt = 'Region admin / Governing System'
                contact_s = g.get('note') or ''
            rows.append(base + [g['cert'], g['status'], channel, nxt, date_s,
                                day_s, time_s, loc_s, link_s, contact_s])
    return rows


def write_worklist(remediations: List[Dict[str, Any]], out_dir: str,
                   stamp: Optional[str] = None,
                   include_compliant: bool = False) -> Dict[str, str]:
    """Write the worklist as both .xlsx and .csv. Returns {'xlsx':..., 'csv':...}."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    stamp = stamp or datetime.now().strftime('%Y%m%d_%H%M%S')
    rows = _worklist_rows(remediations, include_compliant=include_compliant)

    # CSV
    csv_path = out / f'etrainu_compliance_worklist_{stamp}.csv'
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(_WORKLIST_HEADERS)
        w.writerows(rows)

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = 'Worklist'
    ws.append(_WORKLIST_HEADERS)
    head_fill = PatternFill('solid', fgColor='1F3A5F')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = head_fill
        c.alignment = Alignment(vertical='center')
    for row in rows:
        ws.append(row)
    widths = [14, 12, 26, 10, 22, 9, 11, 20, 11, 9, 26, 11, 10, 12, 22, 28, 28]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wdt
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Summary tab
    ws2 = wb.create_sheet('Summary')
    n = len(remediations)
    tracked = [r for r in remediations if r.get('tracked', True)]
    with_gap = sum(1 for r in tracked if r['gaps'])
    unmatched = sum(1 for r in tracked if not r['matched'])
    compliant = sum(1 for r in tracked if r['matched'] and not r['gaps'])
    cert_counts: Dict[str, int] = {}
    chan_counts: Dict[str, int] = {}
    for r in tracked:
        for g in r['gaps']:
            cert_counts[g['cert']] = cert_counts.get(g['cert'], 0) + 1
            chan_counts[g['channel']] = chan_counts.get(g['channel'], 0) + 1
    ws2.append(['Volunteers (all rows)', n])
    ws2.append(['Tracked (Coach/Referee/Youth Ref)', len(tracked)])
    ws2.append(['  Compliant', compliant])
    ws2.append(['  With >=1 gap', with_gap])
    ws2.append(['  Unmatched (verify identity)', unmatched])
    ws2.append([])
    ws2.append(['Gaps by cert', ''])
    for k, val in sorted(cert_counts.items(), key=lambda x: -x[1]):
        ws2.append([k, val])
    ws2.append([])
    ws2.append(['Gaps by channel', ''])
    for k, val in sorted(chan_counts.items(), key=lambda x: -x[1]):
        ws2.append([k, val])
    ws2.column_dimensions['A'].width = 32
    ws2['A1'].font = ws2['A7'].font = Font(bold=True)

    xlsx_path = out / f'etrainu_compliance_worklist_{stamp}.xlsx'
    wb.save(xlsx_path)
    logger.info(f"Worklist written: {xlsx_path} ({len(rows)} rows)")
    return {'xlsx': str(xlsx_path), 'csv': str(csv_path)}


# ──────────────────────────────────────────────────────────────────────────────
#  Output sink #2 — feed the board portal compliance tab
# ──────────────────────────────────────────────────────────────────────────────
def inject_into_portal_payload(payload: Dict[str, Any],
                               remediations: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Mutate a portal compliance.json payload in place, adding to each volunteer:
        next_step  : one-line string for the compliance tab
        next_steps : the structured gap list (for a flyout / detail view)
    Volunteers are matched on (email, division, position) with email-only fallback.
    Returns the same payload."""
    index: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    by_email: Dict[str, Dict[str, Any]] = {}
    for r in remediations:
        key = (r['email'].lower(), r['division'], r['position'])
        index[key] = r
        by_email.setdefault(r['email'].lower(), r)
    for v in payload.get('volunteers', []):
        key = ((v.get('email') or '').lower(), v.get('division', ''), v.get('position', ''))
        r = index.get(key) or by_email.get((v.get('email') or '').lower())
        if not r:
            continue
        v['next_step'] = r['summary']
        v['next_steps'] = r['gaps']
    return payload


def write_portal_next_steps(remediations: List[Dict[str, Any]], out_dir: str,
                            filename: str = 'compliance_next_steps.json') -> str:
    """Standalone next-steps JSON keyed for portal consumption (when you don't
    want to rewrite the whole compliance.json)."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    payload = {
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'volunteers': [
            {'email': r['email'], 'division': r['division'], 'position': r['position'],
             'matched': r['matched'], 'next_step': r['summary'], 'next_steps': r['gaps']}
            for r in remediations
        ],
    }
    path = out / filename
    path.write_text(json.dumps(payload, indent=2, default=str), encoding='utf-8')
    logger.info(f"Portal next-steps written: {path}")
    return str(path)
