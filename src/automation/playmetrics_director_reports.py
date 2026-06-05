"""
PlayMetrics Director Reports

Recreates the four Sports Connect "Division Manager" reports that Division
Coordinators relied on, scoped per division for the PlayMetrics role that
replaces them: the Director (Staff -> Directors). Reports keep the exact legacy
column layouts so Directors get a familiar artifact.

  - enrollment        Enrollment Details  (roster + contact + payment + coach interest)
  - head_coach        Volunteer Verification - Head Coach
  - assistant_coach   Volunteer Verification - Assistant Coach
  - referee           Volunteer Referee

Data sources (downloaded by playmetrics_download_manager into data/playmetrics/):
  - registration-responses_{ts}.csv   -> enrollment report
  - volunteers_{ts}.csv               -> volunteer reports (identity/role/division/contact)
  - coaching-requests_{ts}.csv        -> enriches coach names/phones by email

COMPLIANCE COLUMNS ARE NOT AVAILABLE FROM PLAYMETRICS DURING THE PILOT.
Risk Status/Expiry, Coaching Licence, SafeSport, Fingerprinting, Concussion,
Volunteer Photo, Safe Haven, Sudden Cardiac Arrest live in the Governing System,
which pilot regions cannot access yet (JDP/AYSOU integrations pending). Those
columns are emitted blank by default; set compliance_source to join them from
the existing Sports Affinity / ETrainU pipeline once that bridge is wired.

Column mapping is centralized in FIELD_MAP with three tiers:
  CONFIRMED  - PM column verified against repo code (RegistrarDataContext /
               dashboard generator). Safe.
  INFERRED   - likely PM snake_case name; falls back to blank if absent. Verify
               against a real export header and correct the candidate list.
  UNAVAILABLE- no PM source during pilot (compliance). Always blank.

Usage:
  python main.py --pm-director-reports
  python main.py --pm-director-reports --director-map config/director_map.json
"""
import os
import re
import csv
import glob
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

CONFIRMED, INFERRED, UNAVAILABLE = 'confirmed', 'inferred', 'unavailable'


class PlayMetricsDirectorReports:

    # ── Legacy report column layouts (verbatim from the SC Division Manager samples) ──
    # Core identity/contact columns kept from the legacy layout. The player
    # question answers from Export Responses are appended after these at runtime
    # (see _response_columns), so the report carries the actual registration data.
    ENROLLMENT_COLUMNS = [
        'Program Name', 'Division Name', 'Order Payment Status',
        'Account First Name', 'Account Last Name',
        'Player First Name', 'Player Last Name', 'Player Gender',
        'Player Birth Date', 'Street Address', 'Unit', 'City', 'State',
        'Postal Code', 'User Email', 'Telephone', 'Cellphone', 'Other Phone',
        'Team Name', 'Order Date',
    ]

    # Export Responses source columns already represented by the core columns
    # above (so they are not duplicated when appending question answers).
    _CORE_SOURCE_COLS = {
        'package_name', 'status', 'account_first_name', 'account_last_name',
        'player_first_name', 'player_last_name', 'account_email', 'registered_on',
        'gender', 'birth_date', 'address', 'city', 'state', 'zip', 'account_phone',
    }
    # Internal / duplicate source columns to omit from the appended responses.
    # 'Birth Certificate' dropped per request (verification lives elsewhere).
    _RESPONSE_EXCLUDE = {
        'player_id', 'age_group', 'registered_at', 'package_acct_code', 'Division',
        'Birth Certificate',
    }
    # Team-mate / coach-request questions are only asked in the lower (recreational)
    # divisions; for competitive divisions (10U and up) these columns are omitted.
    # Matched by substring so both the Yes/No question and the free-text answer drop.
    _REQUEST_COL_TOKENS = ('team mate request', 'coach request')
    HEAD_COACH_COLUMNS = [
        'Program Name', 'Division Name', 'User Id', 'Volunteer Id',
        'Volunteer Role', 'Volunteer First Name', 'Volunteer Last Name',
        'Volunteer Email Address', 'Additional First Name', 'Additional Last Name',
        'Secondary Email', 'Volunteer Cellphone', 'Volunteer Telephone',
        'Volunteer Other Phone', 'Risk Status', 'Risk Expiry Date',
        'Coaching Licence', 'SafeSport (Head Coach)',
        'Fingerprinting California Only (Head Coach)',
        'Concussion Awareness (Head Coach)', 'Volunteer Photo (Head Coach)',
        'AYSOs Safe Haven (Head Coach)', 'Sudden Cardiac Arrest (Head Coach)',
        'Associated Participants', 'Coaching Licence Number',
        'Coaching Licence Expiry Date',
    ]
    STANDARD_VOLUNTEER_COLUMNS = [
        'Program Name', 'Division Name', 'Volunteer Role',
        'Volunteer First Name', 'Volunteer Last Name', 'Volunteer Email Address',
        'Volunteer Cellphone', 'Volunteer Telephone', 'Volunteer Other Phone',
        'Risk Status', 'Risk Expiry Date', 'Coaching Licence',
        'SafeSport (Head Coach)', 'Fingerprinting California Only (Head Coach)',
        'Concussion Awareness (Head Coach)', 'Volunteer Photo (Head Coach)',
        'AYSOs Safe Haven (Head Coach)', 'Sudden Cardiac Arrest (Head Coach)',
        'COVID-19 Assumption of Risk, Waiver and Release of Liability Agreement v34536 (Head Coach)',
        'WAIVER, CONSENT, RELEASE, DISCLAIMER AND ASSUMPTION OF RISK AGREEMENT v34535 (Head Coach)',
    ]

    # Compliance columns that the Governing System owns - blank during pilot.
    COMPLIANCE_COLUMNS = {
        'Risk Status', 'Risk Expiry Date', 'Coaching Licence',
        'SafeSport (Head Coach)', 'Fingerprinting California Only (Head Coach)',
        'Concussion Awareness (Head Coach)', 'Volunteer Photo (Head Coach)',
        'AYSOs Safe Haven (Head Coach)', 'Sudden Cardiac Arrest (Head Coach)',
        'Coaching Licence Number', 'Coaching Licence Expiry Date',
        'COVID-19 Assumption of Risk, Waiver and Release of Liability Agreement v34536 (Head Coach)',
        'WAIVER, CONSENT, RELEASE, DISCLAIMER AND ASSUMPTION OF RISK AGREEMENT v34535 (Head Coach)',
    }

    # ── FIELD_MAP: legacy column -> (candidate PM source columns, tier) ──
    # The first present, non-empty candidate wins. Order matters.
    ENROLLMENT_MAP: Dict[str, Any] = {
        # CONFIRMED via RegistrarDataContext.PM_ENROLLMENT_COLS
        'Division Name':        (['package_name'], CONFIRMED),
        'Order Payment Status': (['status'], CONFIRMED),
        'Account First Name':   (['account_first_name'], CONFIRMED),
        'Account Last Name':    (['account_last_name'], CONFIRMED),
        'Player First Name':    (['player_first_name'], CONFIRMED),
        'Player Last Name':     (['player_last_name'], CONFIRMED),
        'User Email':           (['account_email'], CONFIRMED),
        'Order Date':           (['registered_on'], CONFIRMED),
        'Player Gender':        (['gender', 'player_gender'], CONFIRMED),       # M/F -> Male/Female
        'Player Birth Date':    (['birth_date', 'player_birthdate', 'date_of_birth'], CONFIRMED),
        'Street Address':       (['address', 'street', 'street_address'], CONFIRMED),
        'City':                 (['city'], CONFIRMED),
        'State':                (['state'], CONFIRMED),
        'Postal Code':          (['zip', 'postal_code', 'zipcode'], CONFIRMED),
        'Cellphone':            (['account_phone', 'account_mobile', 'cellphone', 'mobile'], CONFIRMED),
        # INFERRED / not present in the Responses export (blank is expected):
        'Program Name':         (['program_name'], INFERRED),                  # else config season
        'Unit':                 (['unit', 'address2', 'apt'], INFERRED),
        'Telephone':            (['telephone', 'home_phone'], INFERRED),       # PM has one phone -> Cellphone
        'Other Phone':          (['other_phone'], INFERRED),
        'Team Name':            (['team', 'team_name'], INFERRED),             # assigned post-draft, not in Responses
    }

    # Volunteer reports draw from volunteers.csv; coach names/phones can be
    # enriched from coaching-requests.csv by email.
    VOLUNTEER_MAP: Dict[str, Any] = {
        'Division Name':            (['package_name'], CONFIRMED),
        'Volunteer Role':           (['volunteer_position'], CONFIRMED),
        'Volunteer Email Address':  (['volunteer_email'], CONFIRMED),
        'Secondary Email':          (['volunteer_secondary_email', 'secondary_email'], INFERRED),
        'Program Name':             (['program_name'], INFERRED),
        'Volunteer First Name':     (['volunteer_first_name', 'first_name'], INFERRED),
        'Volunteer Last Name':      (['volunteer_last_name', 'last_name'], INFERRED),
        'Volunteer Cellphone':      (['volunteer_mobile', 'volunteer_phone', 'mobile', 'phone'], INFERRED),
        'Volunteer Telephone':      (['volunteer_telephone', 'telephone'], INFERRED),
        'Volunteer Other Phone':    (['volunteer_other_phone'], INFERRED),
        'Additional First Name':    (['additional_first_name'], INFERRED),
        'Additional Last Name':     (['additional_last_name'], INFERRED),
        'Associated Participants':  (['associated_participants', 'players', 'player_names'], INFERRED),
        'User Id':                  (['user_id'], INFERRED),
        'Volunteer Id':             (['volunteer_id'], INFERRED),
    }

    POSITION_FILTER = {'head_coach': 'Head Coach', 'assistant_coach': 'Assistant Coach', 'referee': 'Referee'}

    # Merged coach workbook (HC + AC tabs). Sports Connect / Governing System
    # columns are intentionally dropped. Extra volunteers.csv columns are appended
    # after these at runtime so nothing useful is lost.
    COACH_CORE_COLUMNS = [
        'Program Name', 'Division Name', 'Volunteer Role', 'Source',
        'Volunteer First Name', 'Volunteer Last Name', 'Volunteer Email Address',
        'Volunteer Cellphone', 'Volunteer Telephone', 'Secondary Email',
        'Additional First Name', 'Additional Last Name', 'Associated Participants',
    ]
    # volunteers.csv source columns already represented by COACH_CORE_COLUMNS
    # (so they aren't duplicated when appending the extra columns).
    _COACH_SOURCE_COLS = {
        'package_name', 'volunteer_position', 'volunteer_email', 'program_name',
        'volunteer_first_name', 'first_name', 'volunteer_last_name', 'last_name',
        'volunteer_mobile', 'volunteer_phone', 'mobile', 'phone',
        'volunteer_telephone', 'telephone', 'volunteer_other_phone',
        'volunteer_secondary_email', 'secondary_email',
        'additional_first_name', 'additional_last_name',
        'associated_participants', 'players', 'player_names', 'user_id', 'volunteer_id',
    }
    # coaching-requests.csv: candidate columns for the requested role and division.
    _COACH_ROLE_CANDIDATES = ['volunteer_position', 'position', 'requested_position',
                              'coach_role', 'role', 'coaching_role']
    _COACH_DIV_CANDIDATES = ['package_name', 'division', 'age_group', 'team', 'team_name']

    HEADER_FILL = PatternFill('solid', start_color='1C1917')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    BODY_FONT = Font(name='Arial', size=10)
    DATE_COLS = {'Player Birth Date', 'Order Date'}

    def __init__(self, config=None, data_dir: str = None, output_dir: str = None):
        self.config = config
        pm_cfg = (config.get('playmetrics_config', {}) if config else {}) or {}
        self.data_dir = data_dir or pm_cfg.get('download_dir', 'data/playmetrics')
        self.output_dir = Path(output_dir or (Path(self.data_dir) / 'director_reports'))
        self.program_name = (config.get('program_name') if config else None) or pm_cfg.get('program_name')
        self.compliance_source = None  # reserved: future SA/ETrainU join
        # Player-photo embedding (Player Photo column). Photos live behind the
        # PlayMetrics login, so a fetch may need session cookies; on any failure
        # the cell falls back to the original link. Downloaded photos are cached.
        self.embed_photos = (pm_cfg.get('embed_photos', True))
        self.photo_cookies = None        # optional dict/str of PlayMetrics cookies
        self.photo_user_agent = None     # matched from the browser session when prefetching
        self.photo_cache_dir = Path(self.data_dir) / 'player_photos'
        self._photo_session = None

    # =====================================================================
    #  PUBLIC API
    # =====================================================================

    def generate_packets(self, director_map: Dict[str, List[str]] = None,
                         report_types: Iterable[str] = None,
                         responses_file: str = None, volunteers_file: str = None,
                         coaching_file: str = None, output_root: str = None) -> Dict[str, Any]:
        report_types = list(report_types or ('enrollment', 'coaches', 'referee'))
        run_dir = Path(output_root) if output_root else self.output_dir / datetime.now().strftime('%Y%m%d_%H%M%S')
        run_dir.mkdir(parents=True, exist_ok=True)

        built: Dict[str, List[Dict[str, Any]]] = {}
        self._coach_columns = self.COACH_CORE_COLUMNS
        if 'enrollment' in report_types:
            rf = responses_file or self._find_latest('registration-responses')
            if rf:
                built['enrollment'] = self._build_enrollment_rows(self._read_csv(rf))
            else:
                logger.warning("No registration-responses CSV found - skipping enrollment report")
        need_vol = ('coaches' in report_types) or any(rt in self.POSITION_FILTER for rt in report_types)
        if need_vol:
            vf = volunteers_file or self._find_latest('volunteers')
            cf = coaching_file or self._find_latest('coaching-requests')
            if vf:
                vol_rows = self._read_csv(vf)
                coach_rows = self._read_csv(cf) if cf else []
                coach_idx = self._index_coaching(coach_rows)
                if 'coaches' in report_types:
                    records, self._coach_columns = self._build_coach_rows(vol_rows, coach_rows)
                    built['coaches'] = records
                for rt in self.POSITION_FILTER:  # referee (plus legacy hc/ac if explicitly requested)
                    if rt in report_types:
                        built[rt] = self._build_volunteer_rows(vol_rows, rt, coach_idx)
            else:
                logger.warning("No volunteers CSV found - skipping coach/volunteer reports")

        packets = self._resolve_packets(director_map, built)
        manifest = {'run_dir': str(run_dir), 'generated_at': datetime.now().isoformat(),
                    'compliance_note': 'Compliance columns blank: Governing System unavailable to pilot regions.',
                    'packets': {}}

        for pk, divisions in packets.items():
            pdir = run_dir / self._safe(pk)
            pdir.mkdir(parents=True, exist_ok=True)
            files = []
            for rt in report_types:
                if rt == 'coaches':
                    records = built.get('coaches') or []
                    ccols = getattr(self, '_coach_columns', self.COACH_CORE_COLUMNS)
                    for div in divisions:
                        hc = [r for r in records if str(r.get('Division Name', '')) == div
                              and r.get('Volunteer Role') == 'Head Coach']
                        ac = [r for r in records if str(r.get('Division Name', '')) == div
                              and r.get('Volunteer Role') == 'Assistant Coach']
                        if not hc and not ac:
                            continue
                        out = pdir / f"Director_Coaches_{self._safe(div)}.xlsx"
                        self._write_coach_workbook(out, ccols, hc, ac)
                        files.append({'report': 'coaches', 'division': div,
                                      'rows': len(hc) + len(ac), 'path': str(out)})
                    continue
                rows = built.get(rt) or []
                base_cols = self._columns_for(rt)
                for div in divisions:
                    div_rows = [r for r in rows if str(r.get('Division Name', '')) == div]
                    if not div_rows:
                        continue
                    cols = self._columns_for_division(rt, div, base_cols)
                    out = pdir / f"{self._prefix(rt)}_{self._safe(div)}.xlsx"
                    self._write(out, rt, cols, div_rows)
                    files.append({'report': rt, 'division': div, 'rows': len(div_rows), 'path': str(out)})
            manifest['packets'][pk] = {'divisions': divisions, 'files': files}
            logger.info(f"Packet '{pk}': {len(files)} files -> {pdir}")

        mpath = run_dir / 'packet_manifest.json'
        with open(mpath, 'w', encoding='utf-8') as f:
            json.dump(manifest, f, indent=2)
        manifest['manifest_path'] = str(mpath)
        return manifest

    def preview(self, responses_file: str = None, volunteers_file: str = None) -> Dict[str, Any]:
        summary: Dict[str, Any] = {'divisions': {}, 'totals': {}, 'unmapped': {}, 'response_columns': []}
        rf = responses_file or self._find_latest('registration-responses')
        if rf:
            raw = self._read_csv(rf)
            rows = self._build_enrollment_rows(raw)
            summary['totals']['enrollment'] = len(rows)
            summary['response_columns'] = self._response_columns(raw)
            for r in rows:
                summary['divisions'].setdefault(r.get('Division Name', ''), {}).setdefault('enrollment', 0)
                summary['divisions'][r.get('Division Name', '')]['enrollment'] += 1
        vf = volunteers_file or self._find_latest('volunteers')
        if vf:
            raw = self._read_csv(vf)
            summary['unmapped']['volunteers'] = self._unmapped(raw, self.VOLUNTEER_MAP)
            for rt in self.POSITION_FILTER:
                rows = self._build_volunteer_rows(raw, rt, {})
                summary['totals'][rt] = len(rows)
                for r in rows:
                    d = r.get('Division Name', '')
                    summary['divisions'].setdefault(d, {}).setdefault(rt, 0)
                    summary['divisions'][d][rt] += 1
        return summary

    # =====================================================================
    #  ROW BUILDERS
    # =====================================================================

    def _response_columns(self, raw: List[Dict[str, str]]) -> List[str]:
        """Player-question answer columns from Export Responses, in export order,
        minus the columns already represented by the core layout and internal ids."""
        if not raw:
            return []
        return [c for c in raw[0].keys()
                if c not in self._CORE_SOURCE_COLS and c not in self._RESPONSE_EXCLUDE]

    def _build_enrollment_rows(self, raw: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        resp_cols = self._response_columns(raw)
        # Full column order = core identity/contact + the registration responses.
        self._enrollment_columns = list(self.ENROLLMENT_COLUMNS) + resp_cols
        out = []
        for src in raw:
            row = {}
            for col in self.ENROLLMENT_COLUMNS:
                row[col] = self._resolve(src, self.ENROLLMENT_MAP.get(col))
            if not row.get('Program Name'):
                row['Program Name'] = self.program_name or ''
            # Gender: PM exports M/F; legacy report uses Male/Female. Derive if absent.
            row['Player Gender'] = (self._normalize_gender(row.get('Player Gender'))
                                    or self._gender_from_division(row.get('Division Name', '')))
            # Phone: PM gives one E.164 number (+1##########) -> human 10-digit.
            row['Cellphone'] = self._format_phone(row.get('Cellphone', ''))
            # Registration question answers, passed through verbatim.
            for col in resp_cols:
                row[col] = (src.get(col) or '').strip()
            out.append(row)
        out.sort(key=lambda r: (str(r.get('Player Last Name', '')).lower(),
                                str(r.get('Player First Name', '')).lower()))
        return out

    def _build_volunteer_rows(self, raw: List[Dict[str, str]], report_type: str,
                              coach_idx: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
        cols = self._columns_for(report_type)
        target_pos = self.POSITION_FILTER[report_type].casefold()
        out = []
        for src in raw:
            pos = self._resolve(src, self.VOLUNTEER_MAP['Volunteer Role'])
            if pos.casefold() != target_pos:
                continue
            row = {}
            for col in cols:
                if col in self.COMPLIANCE_COLUMNS:
                    row[col] = ''  # Governing System - not available during pilot
                elif col in self.VOLUNTEER_MAP:
                    row[col] = self._resolve(src, self.VOLUNTEER_MAP[col])
                else:
                    row[col] = ''
            if not row.get('Program Name'):
                row['Program Name'] = self.program_name or ''
            # Enrich coach names/phones from coaching-requests by email if blank
            email = (row.get('Volunteer Email Address') or '').lower().strip()
            if email and email in coach_idx:
                cr = coach_idx[email]
                row['Volunteer First Name'] = row.get('Volunteer First Name') or cr.get('coach_first_name', '')
                row['Volunteer Last Name'] = row.get('Volunteer Last Name') or cr.get('coach_last_name', '')
                row['Volunteer Cellphone'] = row.get('Volunteer Cellphone') or cr.get('coach_phone', '')
            out.append(row)
        out.sort(key=lambda r: (str(r.get('Volunteer Last Name', '')).lower(),
                                str(r.get('Volunteer First Name', '')).lower()))
        return out

    @staticmethod
    def _index_coaching(rows: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
        idx = {}
        for r in rows:
            email = (r.get('coach_email') or '').lower().strip()
            if email:
                idx[email] = r
        return idx

    @staticmethod
    def _norm_role(value: str) -> Optional[str]:
        v = (value or '').strip().lower()
        if not v:
            return None
        if 'assistant' in v or v in ('ac', 'asst'):
            return 'Assistant Coach'
        if 'head' in v or v in ('hc', 'coach'):
            return 'Head Coach'
        return None

    def _build_coach_rows(self, vol_rows: List[Dict[str, str]],
                          coach_rows: List[Dict[str, str]]):
        """Union of volunteers.csv (HC/AC) and coaching-requests.csv, keyed by email.
        Returns (records, columns). Each record carries a Source of Volunteer /
        Coaching Request / Both, and any extra volunteers.csv columns pass through."""
        extra_cols = [c for c in (vol_rows[0].keys() if vol_rows else [])
                      if c not in self._COACH_SOURCE_COLS]

        # volunteers side (head/assistant coaches only)
        vol_by_key: Dict[str, Dict[str, Any]] = {}
        for i, src in enumerate(vol_rows):
            role = self._norm_role(self._resolve(src, self.VOLUNTEER_MAP['Volunteer Role']))
            if role not in ('Head Coach', 'Assistant Coach'):
                continue
            email = (self._resolve(src, self.VOLUNTEER_MAP['Volunteer Email Address']) or '').lower().strip()
            vol_by_key[email or f"_vol{i}"] = {'role': role, 'email': email, 'src': src,
                                               'division': self._resolve(src, self.VOLUNTEER_MAP['Division Name'])}

        # coaching-requests side
        coach_by_email: Dict[str, Dict[str, Any]] = {}
        for cr in coach_rows:
            email = (cr.get('coach_email') or '').lower().strip()
            if not email:
                continue
            role = self._norm_role(next((cr[c] for c in self._COACH_ROLE_CANDIDATES
                                         if c in cr and str(cr[c]).strip()), ''))
            division = next((cr[c] for c in self._COACH_DIV_CANDIDATES
                             if c in cr and str(cr[c]).strip()), '')
            coach_by_email[email] = {'role': role, 'division': division, 'cr': cr}

        records = []
        vol_emails = {v['email'] for v in vol_by_key.values() if v['email']}
        keys = set(vol_by_key) | {f"_cr::{e}" for e in coach_by_email if e not in vol_emails}
        for key in keys:
            v = vol_by_key.get(key)
            email = v['email'] if v else key[len('_cr::'):]
            c = coach_by_email.get(email)
            if v is None and c is None:
                continue
            # a volunteer row that also has a coaching request -> attach it
            if v is not None and not c:
                c = coach_by_email.get(v['email'])
            role = (v['role'] if v else None) or (c['role'] if c else None) or 'Head Coach'
            division = (v['division'] if v else '') or (c['division'] if c else '')
            source = 'Both' if (v and c) else ('Volunteer' if v else 'Coaching Request')
            cr = (c['cr'] if c else {})
            src = v['src'] if v else {}
            row = {
                'Program Name': self.program_name or self._resolve(src, self.VOLUNTEER_MAP['Program Name']) or '',
                'Division Name': division,
                'Volunteer Role': role,
                'Source': source,
                'Volunteer First Name': self._resolve(src, self.VOLUNTEER_MAP['Volunteer First Name'])
                                        or cr.get('coach_first_name', ''),
                'Volunteer Last Name': self._resolve(src, self.VOLUNTEER_MAP['Volunteer Last Name'])
                                       or cr.get('coach_last_name', ''),
                'Volunteer Email Address': (v['email'] if v else email),
                'Volunteer Cellphone': self._format_phone(
                    self._resolve(src, self.VOLUNTEER_MAP['Volunteer Cellphone']) or cr.get('coach_phone', '')),
                'Volunteer Telephone': self._resolve(src, self.VOLUNTEER_MAP['Volunteer Telephone']),
                'Secondary Email': self._resolve(src, self.VOLUNTEER_MAP['Secondary Email']),
                'Additional First Name': self._resolve(src, self.VOLUNTEER_MAP['Additional First Name']),
                'Additional Last Name': self._resolve(src, self.VOLUNTEER_MAP['Additional Last Name']),
                'Associated Participants': self._resolve(src, self.VOLUNTEER_MAP['Associated Participants']),
            }
            for col in extra_cols:
                row[col] = (str(src.get(col, '') or '').strip() if src else '')
            records.append(row)

        records.sort(key=lambda r: (str(r.get('Volunteer Last Name', '')).lower(),
                                    str(r.get('Volunteer First Name', '')).lower()))
        return records, (self.COACH_CORE_COLUMNS + extra_cols)

    # =====================================================================
    #  HELPERS
    # =====================================================================

    @staticmethod
    def _resolve(src: Dict[str, str], spec) -> str:
        if not spec:
            return ''
        candidates, _tier = spec
        for cand in candidates:
            if cand in src and str(src[cand]).strip():
                return str(src[cand]).strip()
        return ''

    @staticmethod
    def _normalize_gender(value: str) -> str:
        v = (value or '').strip()
        if v.upper() in ('M', 'MALE', 'BOY'):
            return 'Male'
        if v.upper() in ('F', 'FEMALE', 'GIRL'):
            return 'Female'
        return v

    @staticmethod
    def _format_phone(value: str) -> str:
        """+18185551212 / 18185551212 -> 818-555-1212 (mirrors export manager)."""
        digits = ''.join(ch for ch in str(value or '') if ch.isdigit())
        if len(digits) == 11 and digits.startswith('1'):
            digits = digits[1:]
        if len(digits) == 10:
            return f"{digits[0:3]}-{digits[3:6]}-{digits[6:]}"
        return str(value or '').strip()

    @staticmethod
    def _gender_from_division(division: str) -> str:
        # PM package names like "06UB Boys" / "10UG Girls": 4th char B/G.
        d = (division or '').strip()
        if len(d) >= 4 and d[3].upper() in ('B', 'G'):
            return 'Male' if d[3].upper() == 'B' else 'Female'
        if 'girl' in d.lower():
            return 'Female'
        if 'boy' in d.lower():
            return 'Male'
        return ''

    def _columns_for(self, report_type: str) -> List[str]:
        if report_type == 'enrollment':
            return getattr(self, '_enrollment_columns', self.ENROLLMENT_COLUMNS)
        if report_type == 'head_coach':
            return self.HEAD_COACH_COLUMNS
        return self.STANDARD_VOLUNTEER_COLUMNS

    @staticmethod
    def _is_competitive(division: str) -> bool:
        """True for 10U and up (the competitive divisions). Age = leading number."""
        m = re.match(r'\s*0*(\d{1,2})\s*U', str(division or ''), re.IGNORECASE)
        return bool(m) and int(m.group(1)) >= 10

    def _columns_for_division(self, report_type: str, division: str,
                              base_cols: List[str]) -> List[str]:
        """Drop team-mate / coach-request columns for competitive (10U+) divisions."""
        if report_type != 'enrollment' or not self._is_competitive(division):
            return base_cols
        return [c for c in base_cols
                if not any(tok in c.lower() for tok in self._REQUEST_COL_TOKENS)]

    @staticmethod
    def _prefix(report_type: str) -> str:
        return {'enrollment': 'Director_Enrollment_Details',
                'head_coach': 'Director_Volunteer_Verification_HC',
                'assistant_coach': 'Director_Volunteer_Verification_AC',
                'referee': 'Director_Volunteer_Referee'}[report_type]

    @staticmethod
    def _sheet(report_type: str) -> str:
        return 'Enrollment_Details' if report_type == 'enrollment' else 'Volunteer_Verification_Status'

    def _resolve_packets(self, director_map, built) -> Dict[str, List[str]]:
        if director_map:
            return {k: list(v) for k, v in director_map.items()}
        divisions = set()
        for rows in built.values():
            for r in rows:
                if r.get('Division Name'):
                    divisions.add(str(r['Division Name']))
        return {d: [d] for d in sorted(divisions)}

    @staticmethod
    def _unmapped(raw: List[Dict[str, str]], field_map: Dict[str, Any]) -> List[str]:
        """Source columns present in the export but not mapped to any legacy column."""
        if not raw:
            return []
        mapped = {c for spec in field_map.values() for c in spec[0]}
        return sorted(set(raw[0].keys()) - mapped)

    def _populate_sheet(self, ws, cols: List[str], rows: List[Dict[str, Any]]):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        for r in rows:
            ws.append([self._coerce(r.get(c, ''), c) for c in cols])
        for ri in range(2, ws.max_row + 1):
            for ci, col in enumerate(cols, 1):
                c = ws.cell(row=ri, column=ci)
                c.font = self.BODY_FONT
                if col in self.DATE_COLS and isinstance(c.value, datetime):
                    c.number_format = 'MM/dd/yyyy'
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(max(len(cols),1))}{max(ws.max_row, 1)}"
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = min(max(len(col) + 2, 12), 42)

    def _write(self, out_path: Path, report_type: str, cols: List[str], rows: List[Dict[str, Any]]):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sheet(report_type)
        self._populate_sheet(ws, cols, rows)
        if report_type == 'enrollment' and self.embed_photos and 'Player Photo' in cols:
            self._embed_photos(ws, rows, cols)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)

    def _write_coach_workbook(self, out_path: Path, cols: List[str],
                              hc_rows: List[Dict[str, Any]], ac_rows: List[Dict[str, Any]]):
        """One workbook per division: a Head Coach tab and an Assistant Coach tab."""
        wb = Workbook()
        ws_hc = wb.active
        ws_hc.title = 'HC'
        self._populate_sheet(ws_hc, cols, hc_rows)
        ws_ac = wb.create_sheet('AC')
        self._populate_sheet(ws_ac, cols, ac_rows)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)

    PHOTO_PX = 64  # embedded thumbnail height in pixels

    def _embed_photos(self, ws, rows: List[Dict[str, Any]], cols: List[str]):
        """Replace Player Photo links with embedded thumbnails; fall back to the
        link text for any photo that can't be fetched."""
        from openpyxl.drawing.image import Image as XLImage
        ci = cols.index('Player Photo') + 1
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = max(self.PHOTO_PX / 6.5, 12)
        for ri, r in enumerate(rows, start=2):
            url = (r.get('Player Photo') or '').strip()
            cell = ws.cell(row=ri, column=ci)
            if not url:
                continue
            img_path = self._photo_thumbnail(url)
            if not img_path:
                continue  # leave the link text in the cell as a fallback
            try:
                xi = XLImage(str(img_path))
                # scale to PHOTO_PX tall, keep aspect ratio
                if xi.height:
                    ratio = self.PHOTO_PX / float(xi.height)
                    xi.height = self.PHOTO_PX
                    xi.width = int(xi.width * ratio)
                xi.anchor = f"{letter}{ri}"
                ws.add_image(xi)
                cell.value = None
                ws.row_dimensions[ri].height = self.PHOTO_PX * 0.78
            except Exception as e:
                logger.debug(f"Could not embed photo row {ri}: {e}")

    def _photo_thumbnail(self, url: str) -> Optional[Path]:
        """Return a cached PNG thumbnail for the photo URL, fetching via HTTP if needed."""
        thumb = self._cache_path(url)
        if thumb.exists():
            return thumb
        data = self._fetch_image(url)
        return self._save_thumb_from_bytes(url, data) if data else None

    def _cache_path(self, url: str) -> Path:
        import hashlib
        self.photo_cache_dir.mkdir(parents=True, exist_ok=True)
        return self.photo_cache_dir / (hashlib.md5(url.encode('utf-8')).hexdigest() + '.png')

    def _save_thumb_from_bytes(self, url: str, data: bytes) -> Optional[Path]:
        from io import BytesIO
        try:
            from PIL import Image as PILImage
        except Exception:
            return None
        try:
            im = PILImage.open(BytesIO(data)).convert('RGB')
            im.thumbnail((self.PHOTO_PX * 3, self.PHOTO_PX * 3))
            thumb = self._cache_path(url)
            im.save(thumb, 'PNG')
            return thumb
        except Exception as e:
            logger.debug(f"Bad image data for {url[:40]}...: {e}")
            return None

    def _browser_download(self, driver, url: str, download_dir, timeout: int = 25) -> Optional[bytes]:
        """Navigate the authenticated browser to the photo URL (which serves the
        file as a download) and read the file Chrome drops into download_dir.
        This mirrors opening the URL by hand in the browser."""
        import time
        dl = Path(download_dir)
        try:
            dl.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        before = {p.name for p in dl.iterdir()} if dl.exists() else set()
        try:
            try:
                driver.set_page_load_timeout(8)
            except Exception:
                pass
            try:
                driver.get(url)            # triggers the download; page load usually aborts
            except Exception:
                pass
            deadline = time.time() + timeout
            newfile = None
            while time.time() < deadline:
                cur = [p for p in dl.iterdir()
                       if p.name not in before and not p.name.endswith('.crdownload')]
                if cur:
                    newfile = max(cur, key=lambda p: p.stat().st_mtime)
                    break
                time.sleep(0.4)
            if not newfile:
                self._fetch_diag = (self._fetch_diag or '') + " | browser-download: no file appeared"
                return None
            data = newfile.read_bytes()
            try:
                newfile.unlink()           # don't leave the raw photo in the export dir
            except Exception:
                pass
            return data
        except Exception as e:
            self._fetch_diag = (self._fetch_diag or '') + f" | browser-download error: {e}"
            return None
        finally:
            try:
                driver.set_page_load_timeout(30)
            except Exception:
                pass

    def _fetch_image(self, url: str) -> Optional[bytes]:
        """HTTP GET the photo. Uses PlayMetrics session cookies/UA when available; never raises.
        Records a diagnostic string on self._fetch_diag for troubleshooting."""
        self._fetch_diag = None
        try:
            import requests
        except Exception:
            self._fetch_diag = "requests not installed"
            return None
        if self._photo_session is None:
            self._photo_session = requests.Session()
            if isinstance(self.photo_cookies, dict):
                self._photo_session.cookies.update(self.photo_cookies)
            self._photo_session.headers.update({
                'User-Agent': getattr(self, 'photo_user_agent', None) or
                              'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                'Referer': 'https://playmetrics.com/',
                'Accept': 'image/avif,image/webp,image/png,image/*,*/*;q=0.8',
            })
        try:
            resp = self._photo_session.get(url, timeout=15, allow_redirects=True)
            ct = resp.headers.get('Content-Type', '')
            self._fetch_diag = (f"HTTP {resp.status_code} ct={ct or '?'} "
                                f"len={len(resp.content)} final={resp.url[:60]}")
            # These are attachment/redirect endpoints, so the content-type is often
            # octet-stream/force-download. Accept any 200 with bytes that isn't an
            # HTML page; the image decoder (PIL) does the real validation.
            if resp.status_code == 200 and resp.content and not ct.startswith('text/'):
                return resp.content
        except Exception as e:
            self._fetch_diag = f"requests error: {e}"
        return None

    def prefetch_photos(self, responses_file: str = None,
                        cookies: Dict[str, str] = None) -> Dict[str, int]:
        """Populate the photo cache using an authenticated PlayMetrics session.

        Reuses PlayMetricsDownloadManager for login (so --pm-setup device trust
        applies), lifts the session cookies + User-Agent off the live browser, then
        downloads every Player Photo into data/playmetrics/player_photos/. After
        this runs, report generation embeds photos offline from the cache.

        Pass `cookies` to skip the browser login if you already have a session.
        """
        stats = {'urls': 0, 'cached': 0, 'failed': 0}
        rf = responses_file or self._find_latest('registration-responses')
        if not rf:
            logger.warning("No registration-responses CSV found for photo prefetch")
            return stats
        urls = sorted({(r.get('Player Photo') or '').strip()
                       for r in self._read_csv(rf) if (r.get('Player Photo') or '').strip()})
        stats['urls'] = len(urls)
        if not urls:
            return stats

        mgr = None
        if cookies:
            self.photo_cookies = cookies
        else:
            try:
                from automation.playmetrics_download_manager import PlayMetricsDownloadManager
                mgr = PlayMetricsDownloadManager(config=self.config)
                mgr.initialize()
                if not mgr.login():
                    logger.error("PlayMetrics login failed; cannot fetch photos. Run --pm-setup first.")
                    return stats
                self.photo_cookies = {c['name']: c['value'] for c in mgr.driver.get_cookies()}
                try:
                    self.photo_user_agent = mgr.driver.execute_script("return navigator.userAgent")
                except Exception:
                    self.photo_user_agent = None
            except Exception as e:
                logger.error(f"Could not establish authenticated session for photos: {e}")
                return stats
            finally:
                pass  # keep session open until after downloads; cleaned below

        self._photo_session = None  # rebuild session with the new cookies/UA
        driver = mgr.driver if mgr is not None else None
        dl_dir = getattr(mgr, 'download_dir', self.data_dir) if mgr is not None else self.data_dir
        logged = 0
        try:
            for url in urls:
                if self._cache_path(url).exists():
                    stats['cached'] += 1
                    continue
                # 1) authenticated HTTP GET (follows the redirect to the file)
                data = self._fetch_image(url)
                # 2) fallback: let the browser download it, then read the file
                if not data and driver is not None:
                    data = self._browser_download(driver, url, dl_dir)
                if data and self._save_thumb_from_bytes(url, data):
                    stats['cached'] += 1
                else:
                    stats['failed'] += 1
                    if logged < 5:
                        logger.warning(f"Photo fail [{url[-40:]}]: {getattr(self, '_fetch_diag', '?')}")
                        logged += 1
        finally:
            if mgr is not None:
                try:
                    mgr.cleanup()
                except Exception:
                    pass
        logger.info(f"Photo prefetch: {stats['cached']} cached, {stats['failed']} failed "
                    f"of {stats['urls']} URLs")
        return stats

    def _coerce(self, value, col):
        """Coerce date columns to datetime so Excel formats them; pass through otherwise."""
        if col in self.DATE_COLS and isinstance(value, str) and value.strip():
            v = value.strip()
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S',
                        '%m/%d/%Y %H:%M:%S', '%m/%d/%y'):
                try:
                    return datetime.strptime(v[:19] if 'T' in v or ':' in v else v, fmt)
                except ValueError:
                    continue
        return value

    @staticmethod
    def _read_csv(path: str) -> List[Dict[str, str]]:
        if not path or not os.path.exists(path):
            return []
        with open(path, newline='', encoding='utf-8-sig') as f:
            return list(csv.DictReader(f))

    def _find_latest(self, canonical: str) -> Optional[str]:
        d = Path(self.data_dir)
        if not d.exists():
            return None
        matches = sorted(d.glob(f"{canonical}_*.csv")) or sorted(d.glob(f"{canonical}*.csv"))
        return str(matches[-1]) if matches else None

    @staticmethod
    def _safe(name: str) -> str:
        keep = ''.join(ch if ch.isalnum() or ch in ' -_' else '_' for ch in str(name))
        return keep.strip().replace(' ', '_')[:80]

    @staticmethod
    def format_preview(summary: Dict[str, Any]) -> str:
        lines = ['PlayMetrics Director Reports - Preview', '=' * 52]
        lines.append('Totals: ' + ', '.join(f"{k}={v}" for k, v in summary.get('totals', {}).items()))
        resp = summary.get('response_columns', [])
        if resp:
            lines.append(f"Response columns appended to enrollment ({len(resp)}): " + ', '.join(resp))
        for src, cols in summary.get('unmapped', {}).items():
            if cols:
                lines.append(f"Unmapped {src} columns (candidates to wire): {', '.join(cols)}")
        lines.append('-' * 52)
        for div in sorted(summary.get('divisions', {})):
            detail = ', '.join(f"{k}:{v}" for k, v in summary['divisions'][div].items())
            lines.append(f"  {div}  ({detail})")
        return '\n'.join(lines)


def handle_pm_director_reports(config, args) -> int:
    from utilities.logger import setup_logging
    log = setup_logging(log_level='INFO')
    try:
        rep = PlayMetricsDirectorReports(config=config)
        if getattr(args, 'no_photos', False) or getattr(args, 'director_no_photos', False):
            rep.embed_photos = False
        if getattr(args, 'director_preview', False):
            print(rep.format_preview(rep.preview()))
            return 0
        # Authenticated photo prefetch: logs in, caches photos, then embeds offline.
        if getattr(args, 'director_fetch_photos', False) and rep.embed_photos:
            log.info("Fetching player photos with an authenticated session...")
            rep.prefetch_photos()
        director_map = None
        dm = getattr(args, 'director_map', None)
        if dm and os.path.exists(dm):
            with open(dm, encoding='utf-8') as f:
                director_map = json.load(f)
        manifest = rep.generate_packets(director_map=director_map,
                                        output_root=getattr(args, 'director_output', None))
        n = sum(len(p['files']) for p in manifest['packets'].values())
        print(f"Wrote {n} files across {len(manifest['packets'])} packets -> {manifest['run_dir']}")
        # Upload to Google Drive by default. Disable with --no-upload or by setting
        # director_drive.enabled = false in config. Reports are always saved locally
        # first, so an upload problem never loses the generated files.
        dd = (config.get('director_drive', {}) if config else {}) or {}
        upload = dd.get('enabled', True) and not getattr(args, 'no_upload', False)
        if upload:
            try:
                from integrations.playmetrics_drive_sync import DirectorReportDriveSync
                log.info("Syncing report packets to Google Drive...")
                sync = DirectorReportDriveSync(config=config)
                mpath = os.path.join(manifest['run_dir'], 'packet_manifest.json')
                s = sync.sync_manifest(mpath)
                print(f"Drive sync: {s['uploaded']} uploaded, "
                      f"{s['skipped_no_folder']} skipped (no matching folder), "
                      f"{s['missing_local']} missing local")
            except FileNotFoundError as e:
                print(f"Drive upload skipped (credentials not found): {e}")
            except Exception as e:
                log.error(f"Drive upload failed: {e}")
                print(f"Drive upload failed (reports still saved locally): {e}")
        return 0
    except Exception as e:
        log.error(f"Error generating director reports: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(message)s')
    rep = PlayMetricsDirectorReports()
    print(rep.format_preview(rep.preview()))