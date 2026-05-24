"""
PlayMetrics Enrollment Summary Report — AYSO Region 58
Replaces the Sports Connect / Access database enrollment report.

Reads PM export CSVs and generates a multi-sheet Excel report with:
  1. Enrollment Summary — registrations per division, capacity, percent full
  2. Team Summary — target teams, roster math, allocated/unallocated
  3. Volunteer Summary — coaches, referees, coverage per division
  4. Schedule Summary — team counts, game slots

Usage (standalone):
  python playmetrics_enrollment_report.py

Usage (via main.py):
  python main.py --pm-report
  python main.py --pm-report --pm-report-output data/my_report.xlsx
"""

import os
import math
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# ── Region 58 Division Configuration ────────────────────────────────────

DIVISION_CONFIG = {
    "05U Schoolyard Coed": {"max_spots": 60, "roster_size": 6,  "roster_min": 5,  "on_field": 4,  "refs_required": False, "sort": 1},
    "06UB Boys":           {"max_spots": 60, "roster_size": 6,  "roster_min": 5,  "on_field": 4,  "refs_required": False, "sort": 2},
    "06UG Girls":          {"max_spots": 36, "roster_size": 6,  "roster_min": 5,  "on_field": 4,  "refs_required": False, "sort": 3},
    "07UB Boys":           {"max_spots": 84, "roster_size": 8,  "roster_min": 6,  "on_field": 5,  "refs_required": False, "sort": 4},
    "07UG Girls":          {"max_spots": 42, "roster_size": 7,  "roster_min": 6,  "on_field": 4,  "refs_required": False, "sort": 5},
    "08UB Boys":           {"max_spots": 70, "roster_size": 7,  "roster_min": 6,  "on_field": 5,  "refs_required": False, "sort": 6},
    "08UG Girls":          {"max_spots": 56, "roster_size": 7,  "roster_min": 6,  "on_field": 5,  "refs_required": False, "sort": 7},
    "10UB Boys":           {"max_spots": 180, "roster_size": 9,  "roster_min": 8,  "on_field": 7,  "refs_required": True,  "sort": 8},
    "10UG Girls":          {"max_spots": 126, "roster_size": 9,  "roster_min": 8,  "on_field": 7,  "refs_required": True,  "sort": 9},
    "12UB Boys":           {"max_spots": 144, "roster_size": 12, "roster_min": 10, "on_field": 9,  "refs_required": True,  "sort": 10},
    "12UG Girls":          {"max_spots": 72,  "roster_size": 12, "roster_min": 10, "on_field": 9,  "refs_required": True,  "sort": 11},
    "14UB Boys":           {"max_spots": 84,  "roster_size": 14, "roster_min": 12, "on_field": 11, "refs_required": True,  "sort": 12},
    "14UG Girls":          {"max_spots": 28,  "roster_size": 14, "roster_min": 12, "on_field": 11, "refs_required": True,  "sort": 13},
    "16UB Boys":           {"max_spots": 48,  "roster_size": 14, "roster_min": 12, "on_field": 11, "refs_required": True,  "sort": 14},
    "16UG Girls":          {"max_spots": 20,  "roster_size": 14, "roster_min": 12, "on_field": 11, "refs_required": True,  "sort": 15},
    "19UB Boys":           {"max_spots": 22,  "roster_size": 22, "roster_min": 12, "on_field": 11, "refs_required": True,  "sort": 16},
    "19UG Girls":          {"max_spots": 22,  "roster_size": 22, "roster_min": 12, "on_field": 11, "refs_required": True,  "sort": 17},
}

# Map age_group values from exports to division names for volunteer matching
AGE_GROUP_TO_DIVISIONS = {
    "5U": ["05U Schoolyard Coed"],
    "6U": ["06UB Boys", "06UG Girls"],
    "7U": ["07UB Boys", "07UG Girls"],
    "8U": ["08UB Boys", "08UG Girls"],
    "10U": ["10UB Boys", "10UG Girls"],
    "12U": ["12UB Boys", "12UG Girls"],
    "14U": ["14UB Boys", "14UG Girls"],
    "16U": ["16UB Boys", "16UG Girls"],
    "19U": ["19UB Boys", "19UG Girls"],
}


class PlayMetricsEnrollmentReport:
    """Generates enrollment summary report from PlayMetrics export CSVs."""

    def __init__(self, data_dir: str = "data/playmetrics"):
        self.data_dir = Path(data_dir)
        self.responses_df: Optional[pd.DataFrame] = None
        self.volunteers_df: Optional[pd.DataFrame] = None
        self.coaching_df: Optional[pd.DataFrame] = None

    def _find_latest(self, pattern: str) -> Optional[Path]:
        candidates = list(self.data_dir.glob(pattern))
        if not candidates:
            return None
        return max(candidates, key=lambda p: p.stat().st_mtime)

    def load_data(self):
        """Load all PM export CSVs."""
        # Registration responses
        for pattern in ["registration-responses*.csv", "registration_responses*.csv"]:
            path = self._find_latest(pattern)
            if path:
                self.responses_df = pd.read_csv(path, encoding="utf-8")
                logger.info(f"Loaded responses: {path} ({len(self.responses_df)} rows)")
                break
        if self.responses_df is None:
            logger.warning("No registration responses CSV found")

        # Volunteers
        path = self._find_latest("volunteers*.csv")
        if path:
            self.volunteers_df = pd.read_csv(path, encoding="utf-8")
            logger.info(f"Loaded volunteers: {path} ({len(self.volunteers_df)} rows)")

        # Coaching requests
        path = self._find_latest("*coaching-requests*.csv") or self._find_latest("*coaching_requests*.csv")
        if path:
            self.coaching_df = pd.read_csv(path, encoding="utf-8")
            logger.info(f"Loaded coaching requests: {path} ({len(self.coaching_df)} rows)")

    # ── Section 1: Enrollment Summary ───────────────────────────────────

    def build_enrollment_summary(self) -> pd.DataFrame:
        """Build enrollment summary by division."""
        rows = []
        for div_name, config in sorted(DIVISION_CONFIG.items(), key=lambda x: x[1]["sort"]):
            enrolled = 0
            waitlisted = 0
            if self.responses_df is not None:
                div_df = self.responses_df[self.responses_df["package_name"] == div_name]
                enrolled = len(div_df[div_df["status"] == "Completed"])
                waitlisted = len(div_df[div_df["status"] == "Waitlisted"]) if "Waitlisted" in div_df["status"].values else 0

            max_spots = config["max_spots"]
            pct_full = round(enrolled / max_spots * 100, 1) if max_spots else 0
            remaining = max(0, max_spots - enrolled)

            rows.append({
                "Division": div_name,
                "Enrolled": enrolled,
                "Max Spots": max_spots,
                "Remaining": remaining,
                "Waitlist": waitlisted,
                "% Full": pct_full,
            })

        # Totals row
        df = pd.DataFrame(rows)
        totals = {
            "Division": "TOTAL",
            "Enrolled": df["Enrolled"].sum(),
            "Max Spots": df["Max Spots"].sum(),
            "Remaining": df["Remaining"].sum(),
            "Waitlist": df["Waitlist"].sum(),
            "% Full": round(df["Enrolled"].sum() / df["Max Spots"].sum() * 100, 1) if df["Max Spots"].sum() else 0,
        }
        df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
        return df

    # ── Section 2: Team Summary ─────────────────────────────────────────

    def build_team_summary(self) -> pd.DataFrame:
        """Build team formation summary by division."""
        rows = []
        for div_name, config in sorted(DIVISION_CONFIG.items(), key=lambda x: x[1]["sort"]):
            enrolled = 0
            if self.responses_df is not None:
                enrolled = len(self.responses_df[
                    (self.responses_df["package_name"] == div_name) &
                    (self.responses_df["status"] == "Completed")
                ])

            roster_size = config["roster_size"]
            on_field = config["on_field"]
            target_teams = math.ceil(enrolled / roster_size) if enrolled > 0 else 0
            allocated = target_teams * roster_size
            unallocated = enrolled - (target_teams * roster_size) if enrolled > 0 else 0
            actual_roster = round(enrolled / target_teams, 1) if target_teams > 0 else 0

            rows.append({
                "Division": div_name,
                "Enrolled": enrolled,
                "Roster Size": roster_size,
                "On-Field": on_field,
                "Target Teams": target_teams,
                "Actual Roster": actual_roster,
                "Subs": round(actual_roster - on_field, 1) if target_teams > 0 else 0,
            })

        df = pd.DataFrame(rows)
        totals = {
            "Division": "TOTAL",
            "Enrolled": df["Enrolled"].sum(),
            "Roster Size": "",
            "On-Field": "",
            "Target Teams": df["Target Teams"].sum(),
            "Actual Roster": "",
            "Subs": "",
        }
        df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
        return df

    # ── Section 3: Volunteer Summary ────────────────────────────────────

    def build_volunteer_summary(self) -> pd.DataFrame:
        """Build volunteer coverage summary by division."""
        # Count coaching requests per division
        coach_counts = {}
        if self.coaching_df is not None:
            for _, row in self.coaching_df.iterrows():
                # Get the player's age group to determine division
                player_id = row.get("player_id", "")
                # Look up the player's package from responses
                if self.responses_df is not None and player_id:
                    player_match = self.responses_df[self.responses_df["player_id"] == player_id]
                    if not player_match.empty:
                        div_name = player_match.iloc[0].get("package_name", "")
                        if div_name not in coach_counts:
                            coach_counts[div_name] = {"head": 0, "asst": 0}
                        if str(row.get("request_head_coach", "")).upper() == "Y":
                            coach_counts[div_name]["head"] += 1
                        if str(row.get("request_asst_coach", "")).upper() == "Y":
                            coach_counts[div_name]["asst"] += 1

        # Count volunteers by position and division
        vol_counts = {}
        if self.volunteers_df is not None:
            for _, row in self.volunteers_df.iterrows():
                div_name = str(row.get("package_name", ""))
                position = str(row.get("volunteer_position", ""))
                if div_name not in vol_counts:
                    vol_counts[div_name] = {"Head Coach": 0, "Assistant Coach": 0, "Referee": 0, "Team Manager": 0, "Board Member": 0}
                if position in vol_counts[div_name]:
                    vol_counts[div_name][position] += 1

        rows = []
        for div_name, config in sorted(DIVISION_CONFIG.items(), key=lambda x: x[1]["sort"]):
            enrolled = 0
            if self.responses_df is not None:
                enrolled = len(self.responses_df[
                    (self.responses_df["package_name"] == div_name) &
                    (self.responses_df["status"] == "Completed")
                ])

            roster_size = config["roster_size"]
            target_teams = math.ceil(enrolled / roster_size) if enrolled > 0 else 0

            # Coaching requests (from coaching requests export)
            cr = coach_counts.get(div_name, {"head": 0, "asst": 0})
            # Volunteer signups (from volunteer export)
            vc = vol_counts.get(div_name, {"Head Coach": 0, "Assistant Coach": 0, "Referee": 0, "Team Manager": 0, "Board Member": 0})

            # Combine both sources for head coaches
            total_hc = cr["head"] + vc.get("Head Coach", 0)
            total_ac = cr["asst"] + vc.get("Assistant Coach", 0)

            hc_pct = round(total_hc / target_teams * 100, 1) if target_teams > 0 else 0

            rows.append({
                "Division": div_name,
                "Target Teams": target_teams,
                "HC Requests": total_hc,
                "HC Coverage": f"{hc_pct}%",
                "AC Requests": total_ac,
                "Referees": vc.get("Referee", 0),
                "Team Mgrs": vc.get("Team Manager", 0),
            })

        df = pd.DataFrame(rows)
        totals = {
            "Division": "TOTAL",
            "Target Teams": df["Target Teams"].sum(),
            "HC Requests": df["HC Requests"].sum(),
            "HC Coverage": f"{round(df['HC Requests'].sum() / df['Target Teams'].sum() * 100, 1)}%" if df["Target Teams"].sum() > 0 else "0%",
            "AC Requests": df["AC Requests"].sum(),
            "Referees": df["Referees"].sum(),
            "Team Mgrs": df["Team Mgrs"].sum(),
        }
        df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
        return df

    # ── Section 4: Schedule Summary ─────────────────────────────────────

    def build_schedule_summary(self) -> pd.DataFrame:
        """Build schedule capacity summary."""
        rows = []
        for div_name, config in sorted(DIVISION_CONFIG.items(), key=lambda x: x[1]["sort"]):
            enrolled = 0
            if self.responses_df is not None:
                enrolled = len(self.responses_df[
                    (self.responses_df["package_name"] == div_name) &
                    (self.responses_df["status"] == "Completed")
                ])

            roster_size = config["roster_size"]
            target_teams = math.ceil(enrolled / roster_size) if enrolled > 0 else 0
            # Games per Saturday = teams / 2 (each game has 2 teams)
            games_per_week = math.ceil(target_teams / 2) if target_teams > 0 else 0

            rows.append({
                "Division": div_name,
                "Teams": target_teams,
                "Games/Week": games_per_week,
            })

        df = pd.DataFrame(rows)
        totals = {
            "Division": "TOTAL",
            "Teams": df["Teams"].sum(),
            "Games/Week": df["Games/Week"].sum(),
        }
        df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
        return df

    # ── Report Generation ───────────────────────────────────────────────

    def generate_report(self, output_path: str = None) -> str:
        """Generate the enrollment summary report matching the original SC/Access format.
        
        Single sheet with all four sections side by side:
          Cols A-I:   Enrollment Summary
          Cols J-K:   Roster Config  
          Cols L-U:   Team Summary
          Cols V-AC:  Volunteer Summary
          Cols AD-AF: Schedule Summary
        """
        self.load_data()

        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = f"data/Enrollment_Summary_Report_{timestamp}.xlsx"

        enrollment = self.build_enrollment_summary()
        teams = self.build_team_summary()
        volunteers = self.build_volunteer_summary()
        schedule = self.build_schedule_summary()

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.formatting.rule import ColorScaleRule

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enrollment_Summary_Report"

        # Styles matching original
        calibri = Font(name="Calibri", size=11)
        header_font = Font(name="Calibri", size=11, bold=True)
        section_font = Font(name="Calibri", size=11, bold=True)
        
        # Original uses theme 3 tint 0.6 — #ACB9CA across entire sheet
        data_fill = PatternFill("solid", fgColor="ACB9CA")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Base fill: apply #ACB9CA to entire sheet area
        max_row = len(DIVISION_CONFIG) + 30  # extend well beyond data rows
        for row in range(1, max_row + 1):
            for col in range(1, 33):
                cell = ws.cell(row=row, column=col)
                cell.fill = data_fill
                cell.font = calibri

        # Row 1: Section headers (bold, centered — on top of base fill)
        sections = {
            "C": "Enrollment Summary",
            "L": "Team Summary", 
            "V": "Volunteer Summary",
            "AD": "Schedule Summary",
        }
        for col, title in sections.items():
            cell = ws[f"{col}1"]
            cell.value = title
            cell.font = section_font
            cell.alignment = Alignment(horizontal="center")

        # Row 2: Column headers (bold, centered — on top of base fill)
        headers = [
            "Program Name", "Division Name", "Reg Close",
            "Division Enrollments", "Maximum", "Waitlist", "Percent",
            "Unpaid", "Percent",
            "Roster Size", "On Field",
            "Current Teams", "Target Teams", "Percent",
            "Teams Formed", "Allocated", "Unallocated",
            "Paid Unallocated", "Unpaid Unallocated",
            "Other Teams", "Waitlist Teams",
            "Head Coach", "Percent Head Coach", "Assistant Coach",
            "Referees Needed", "AR Needed", "Total Referees",
            "Referees", "Youth Referees",
            "Team Count", "Team Max", "Game Max",
        ]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font = header_font
            cell.alignment = center_align

        # Row 3: Totals
        ws.cell(row=3, column=1, value="2026 Fall Core")
        ws.cell(row=3, column=2, value="Totals")

        # Data rows starting at row 4
        row_num = 4
        for div_name in sorted(DIVISION_CONFIG.keys(), key=lambda x: DIVISION_CONFIG[x]["sort"]):
            config = DIVISION_CONFIG[div_name]
            
            # Enrollment data
            enrolled = 0
            waitlisted = 0
            if self.responses_df is not None:
                div_df = self.responses_df[self.responses_df["package_name"] == div_name]
                enrolled = len(div_df[div_df["status"] == "Completed"])
                if "Waitlisted" in div_df["status"].values:
                    waitlisted = len(div_df[div_df["status"] == "Waitlisted"])

            max_spots = config["max_spots"]
            roster_size = config["roster_size"]
            on_field = config["on_field"]
            pct_full = enrolled / max_spots if max_spots else 0
            target_teams = math.ceil(enrolled / roster_size) if enrolled > 0 else 0
            allocated = target_teams * roster_size if target_teams > 0 else 0
            unallocated = enrolled - allocated if enrolled > 0 else 0
            pct_teams = 0  # Current teams / target — 0 until teams are formed
            games_per_week = math.ceil(target_teams / 2) if target_teams > 0 else 0

            # Volunteer counts
            hc = volunteers[volunteers["Division"] == div_name]["HC Requests"].values
            hc_count = int(hc[0]) if len(hc) > 0 else 0
            ac = volunteers[volunteers["Division"] == div_name]["AC Requests"].values
            ac_count = int(ac[0]) if len(ac) > 0 else 0
            refs = volunteers[volunteers["Division"] == div_name]["Referees"].values
            ref_count = int(refs[0]) if len(refs) > 0 else 0
            hc_pct = hc_count / target_teams if target_teams > 0 else 0

            # Refs needed: 1 center + 2 AR per game for competitive divisions
            is_competitive = config["refs_required"]
            refs_needed = target_teams // 2 if is_competitive else 0
            ar_needed = refs_needed * 2 if is_competitive else 0

            data = [
                "",  # A: Program Name (blank after first)
                div_name,  # B: Division
                None,  # C: Reg Close
                enrolled,  # D: Enrollments
                max_spots,  # E: Maximum
                waitlisted if waitlisted else None,  # F: Waitlist
                pct_full,  # G: % Full
                None,  # H: Unpaid
                None,  # I: % Unpaid
                roster_size,  # J: Roster Size
                on_field,  # K: On Field
                0,  # L: Current Teams (none formed yet)
                target_teams,  # M: Target Teams
                pct_teams,  # N: % Teams
                0,  # O: Teams Formed
                allocated if enrolled > 0 else None,  # P: Allocated
                unallocated if unallocated != 0 else 0,  # Q: Unallocated
                None,  # R: Paid Unallocated
                None,  # S: Unpaid Unallocated
                None,  # T: Other Teams
                None,  # U: WL Teams
                hc_count if hc_count else None,  # V: Head Coach
                hc_pct if target_teams > 0 else None,  # W: % HC
                ac_count if ac_count else None,  # X: Asst Coach
                refs_needed if refs_needed else None,  # Y: Refs Needed
                ar_needed if ar_needed else None,  # Z: AR Needed
                ref_count if ref_count else None,  # AA: Total Refs
                ref_count if ref_count else None,  # AB: Referees
                None,  # AC: Youth Refs
                target_teams if is_competitive else None,  # AD: Team Count
                max_spots // roster_size if is_competitive else None,  # AE: Team Max
                games_per_week if is_competitive else None,  # AF: Game Max
            ]

            for i, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=i, value=val)
                cell.border = thin_border
                # Format percentages as 0.0%
                if i in (7, 9, 14, 23) and isinstance(val, (int, float)):
                    cell.number_format = '0.0%'
                # Format integers
                elif i in (4, 5, 30, 31, 32) and isinstance(val, (int, float)):
                    cell.number_format = '0'

            row_num += 1

        # Row 3: Totals with SUM formulas
        last_row = row_num - 1
        sum_cols = {
            4: f"=SUM(D4:D{last_row})",   # Enrollments
            5: f"=SUM(E4:E{last_row})",   # Maximum
            6: f"=SUM(F4:F{last_row})",   # Waitlist
            7: f"=D3/E3",                  # % Full
            12: f"=SUM(L4:L{last_row})",  # Current Teams
            13: f"=SUM(M4:M{last_row})",  # Target Teams
            14: f"=IF(M3=0,0,L3/M3)",     # % Teams
            15: f"=SUM(O4:O{last_row})",  # Teams Formed
            16: f"=SUM(P4:P{last_row})",  # Allocated
            17: f"=SUM(Q4:Q{last_row})",  # Unallocated
            22: f"=SUM(V4:V{last_row})",  # Head Coach
            23: f"=IF(M3=0,0,IF(V3/M3>1,1,V3/M3))",  # % HC
            24: f"=SUM(X4:X{last_row})",  # Asst Coach
            27: f"=SUM(AA4:AA{last_row})", # Total Refs
            28: f"=SUM(AB4:AB{last_row})", # Referees
        }
        for col, formula in sum_cols.items():
            cell = ws.cell(row=3, column=col, value=formula)
            if col in (7, 14, 23):
                cell.number_format = '0.0%'

        # Color scale conditional formatting (gradient red → yellow → green)
        # Matches original: colorScale on % columns
        pct_ranges = [
            f"G3:G{last_row}",   # % Full
            f"I3:I{last_row}",   # % Unpaid
            f"N3:N{last_row}",   # % Teams
            f"W3:W{last_row}",   # % Head Coach
        ]
        for rng in pct_ranges:
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min", start_color="F8696B",   # Red
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",  # Yellow
                    end_type="max", end_color="63BE7B",       # Green
                )
            )

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        for col_letter in "CDEFGHIJKLMNOPQRSTUVWXYZ":
            ws.column_dimensions[col_letter].width = 11
        ws.column_dimensions["AA"].width = 11
        ws.column_dimensions["AB"].width = 11
        ws.column_dimensions["AC"].width = 11
        ws.column_dimensions["AD"].width = 11
        ws.column_dimensions["AE"].width = 11
        ws.column_dimensions["AF"].width = 11

        # Freeze panes at row 3 col B
        ws.freeze_panes = "C4"

        wb.save(output_path)
        logger.info(f"Report saved to: {output_path}")
        return output_path

    def print_console_summary(self):
        """Print a formatted summary to console."""
        self.load_data()

        enrollment = self.build_enrollment_summary()
        teams = self.build_team_summary()
        volunteers = self.build_volunteer_summary()

        print("\n" + "=" * 70)
        print(f"  AYSO Region 58 — Enrollment Summary Report")
        print(f"  Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
        print("=" * 70)

        print("\n📊 ENROLLMENT SUMMARY")
        print("-" * 60)
        print(f"{'Division':<25} {'Enrolled':>8} {'Max':>6} {'Remain':>8} {'% Full':>8}")
        print("-" * 60)
        for _, row in enrollment.iterrows():
            div = row["Division"]
            pct = row["% Full"]
            bar = "█" * int(pct / 5) if div != "TOTAL" else ""
            print(f"{div:<25} {row['Enrolled']:>8} {row['Max Spots']:>6} {row['Remaining']:>8} {pct:>7}% {bar}")

        print(f"\n⚽ TEAM SUMMARY")
        print("-" * 60)
        print(f"{'Division':<25} {'Enrolled':>8} {'Teams':>6} {'Roster':>8} {'Subs':>6}")
        print("-" * 60)
        for _, row in teams.iterrows():
            print(f"{row['Division']:<25} {str(row['Enrolled']):>8} {str(row['Target Teams']):>6} {str(row['Actual Roster']):>8} {str(row['Subs']):>6}")

        print(f"\n🙋 VOLUNTEER SUMMARY")
        print("-" * 60)
        print(f"{'Division':<25} {'Teams':>6} {'HC':>5} {'Cvg':>6} {'AC':>5} {'Ref':>5}")
        print("-" * 60)
        for _, row in volunteers.iterrows():
            print(f"{row['Division']:<25} {str(row['Target Teams']):>6} {str(row['HC Requests']):>5} {str(row['HC Coverage']):>6} {str(row['AC Requests']):>5} {str(row['Referees']):>5}")

        # Quick stats
        total_row = enrollment[enrollment["Division"] == "TOTAL"].iloc[0]
        print(f"\n{'─' * 40}")
        print(f"Total Enrolled: {int(total_row['Enrolled'])} / {int(total_row['Max Spots'])} ({total_row['% Full']}%)")
        total_teams = teams[teams["Division"] == "TOTAL"].iloc[0]["Target Teams"]
        print(f"Total Teams: {int(total_teams)}")
        total_vol = volunteers[volunteers["Division"] == "TOTAL"].iloc[0]
        print(f"Head Coaches: {int(total_vol['HC Requests'])} / {int(total_vol['Target Teams'])} needed ({total_vol['HC Coverage']})")
        print()


# ── CLI Entry Points ────────────────────────────────────────────────────

def handle_pm_report(config, args) -> int:
    """Handle PM enrollment report from command line."""
    try:
        data_dir = getattr(args, "pm_data_dir", "data/playmetrics")
        output = getattr(args, "pm_report_output", None)

        report = PlayMetricsEnrollmentReport(data_dir=data_dir)
        report.print_console_summary()

        output_path = report.generate_report(output_path=output)
        print(f"📄 Excel report saved to: {output_path}")
        return 0
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    report = PlayMetricsEnrollmentReport()
    report.print_console_summary()
    path = report.generate_report()
    print(f"\nExcel report saved to: {path}")