"""
Game Card Processor for AYSO League Game Cards
Fills upper division details from Enrollment Details report
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import shutil
import glob

logger = logging.getLogger(__name__)


class GameCardProcessor:
    """Process league game cards by filling upper division details from Enrollment Details"""
    
    UPPER_DIVISIONS = ['16UG', '16UB', '19UG', '19UB']
    
    # Column mappings for data extraction
    ENROLLMENT_COLUMNS = {
        'player_first': 'Player First Name',
        'player_last': 'Player Last Name',
        'player_id': 'Player ID',
        'division': 'Division Name',
        'team': 'Team Name',
        'team_number': 'Team Number',
        'jersey': 'Player Jersey Number',
        'birth_date': 'Player Birth Date',
        'order_no': 'Order No',
        'payment_status': 'Order Payment Status',
        'parent_first': 'Account First Name',
        'parent_last': 'Account Last Name',
        'parent_email': 'User Email',
        'registration_status': 'Registration Status'
    }
    
    def __init__(self, config):
        """
        Initialize Game Card Processor
        
        Args:
            config: Configuration manager instance
        """
        self.config = config
        self.download_dir = config.get('download_dir', 'data/downloads')
        self.output_dir = config.get('game_card_output_dir', 'data/game_cards')
        
        # Create output directory if it doesn't exist
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        game_card_config = config.get('game_card_config', {})
        self.upper_divisions = game_card_config.get('upper_divisions', self.UPPER_DIVISIONS)
        self.backup_original = game_card_config.get('backup_original', True)
        self.output_suffix = game_card_config.get('output_suffix', '_processed')
        
        # Coach data cache (would be populated from volunteer report)
        self.coach_data = {}
        
        # Division name mapping (actual name -> standard name)
        self.division_mapping = {}
        
    def find_latest_enrollment_file(self) -> Optional[str]:
        """Find the latest Enrollment Details file"""
        pattern = os.path.join(self.download_dir, "Enrollment_Details*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            logger.error("No Enrollment Details file found")
            return None
            
        # Get the most recent file
        latest_file = max(files, key=os.path.getmtime)
        logger.info(f"Found enrollment file: {latest_file}")
        return latest_file
        
    def find_latest_volunteer_file(self) -> Optional[str]:
        """Find the latest Volunteer Details file for coach information"""
        pattern = os.path.join(self.download_dir, "Volunteer_Details*.xlsx")
        files = glob.glob(pattern)
        
        if files:
            return max(files, key=os.path.getmtime)
        return None
        
    def load_enrollment_data(self, file_path: str = None) -> pd.DataFrame:
        """
        Load and prepare enrollment data
        
        Args:
            file_path: Path to enrollment file (auto-finds if None)
            
        Returns:
            Prepared DataFrame with enrollment data
        """
        if not file_path:
            file_path = self.find_latest_enrollment_file()
            
        if not file_path:
            raise FileNotFoundError("No enrollment file found")
            
        logger.info(f"Loading enrollment data from: {file_path}")
        df = pd.read_excel(file_path)
        
        # Standardize column names
        df.columns = df.columns.str.strip()
        
        # Debug: Show unique division names
        unique_divisions = df['Division Name'].unique()
        logger.info(f"Unique divisions in file: {unique_divisions}")
        
        # Build division mapping
        self._build_division_mapping(unique_divisions)
        
        # Filter for upper divisions - check if any of our target divisions are contained in the division name
        df_upper = df[df['Division Name'].apply(lambda x: self._is_upper_division(str(x)))].copy()
        
        # Add standardized division column for easier processing
        df_upper['Standard Division'] = df_upper['Division Name'].apply(
            lambda x: self.division_mapping.get(str(x), str(x))
        )
        
        # Filter out players who are not on a team (Unallocated)
        if 'Team Name' in df_upper.columns:
            df_upper = df_upper[df_upper['Team Name'] != 'Unallocated']
            df_upper = df_upper[df_upper['Team Name'].notna()]  # Also remove null team names
        
        logger.info(f"Found {len(df_upper)} players in upper divisions with team assignments")
        logger.info(f"Division breakdown: {df_upper['Standard Division'].value_counts().to_dict()}")
        
        return df_upper
        
    def _build_division_mapping(self, division_names):
        """
        Build a mapping from actual division names to standard names
        
        Args:
            division_names: List of division names from the Excel file
        """
        self.division_mapping = {}
        
        for div_name in division_names:
            if pd.isna(div_name):
                continue
                
            div_str = str(div_name)
            for target_div in self.upper_divisions:
                if self._matches_target_division(div_str, target_div):
                    self.division_mapping[div_str] = target_div
                    logger.debug(f"Mapped '{div_str}' to '{target_div}'")
                    break
                    
    def _is_upper_division(self, division_name: str) -> bool:
        """
        Check if a division name represents an upper division
        
        Args:
            division_name: The division name from the Excel file
            
        Returns:
            True if this is an upper division
        """
        # Normalize the division name
        normalized = division_name.upper().replace(' ', '').replace('-', '')
        
        # Check each target division
        for target_div in self.upper_divisions:
            # Extract age and gender from target (e.g., "16UG" -> "16" and "G")
            age = target_div[:2]
            gender = target_div[-1]
            
            # Various possible formats
            patterns = [
                f"{age}U{gender}",           # 16UG
                f"{age}U{gender.lower()}",   # 16Ug
                f"U{age}{gender}",           # U16G
                f"U{age}{gender.lower()}",   # U16g
                f"{age}{gender}",            # 16G
                f"{age}U",                   # 16U (if gender specified elsewhere)
            ]
            
            # Also check for full gender names
            if gender == 'G':
                patterns.extend([
                    f"{age}UGIRLS", f"{age}UGIRL", f"U{age}GIRLS", f"U{age}GIRL",
                    f"{age}GIRLS", f"{age}GIRL"
                ])
            elif gender == 'B':
                patterns.extend([
                    f"{age}UBOYS", f"{age}UBOY", f"U{age}BOYS", f"U{age}BOY",
                    f"{age}BOYS", f"{age}BOY"
                ])
            
            # Check if any pattern matches
            for pattern in patterns:
                if pattern in normalized:
                    logger.debug(f"Division '{division_name}' matched as {target_div}")
                    return True
                    
        return False
        
    def load_coach_data(self, file_path: str = None) -> Dict[str, Dict]:
        """
        Load coach data from volunteer report
        
        Args:
            file_path: Path to volunteer file (auto-finds if None)
            
        Returns:
            Dictionary mapping team identifiers to coach info
        """
        if not file_path:
            file_path = self.find_latest_volunteer_file()
            
        if not file_path:
            logger.warning("No volunteer file found - coach data will be empty")
            return {}
            
        try:
            logger.info(f"Loading coach data from: {file_path}")
            df = pd.read_excel(file_path)
            
            # Filter for coaches and assistant coaches
            coach_df = df[df['Volunteer Role'].str.contains('Coach', case=False, na=False)]
            
            coach_data = {}
            for _, row in coach_df.iterrows():
                division = row.get('Division Name', '')
                team = row.get('Team Name', '')
                
                if division and team:
                    key = f"{division}_{team}"
                    role = row.get('Volunteer Role', '').lower()
                    
                    if key not in coach_data:
                        coach_data[key] = {}
                    
                    if 'head' in role or role == 'coach':
                        coach_data[key]['head_coach'] = {
                            'name': f"{row.get('First Name', '')} {row.get('Last Name', '')}",
                            'email': row.get('Email', ''),
                            'phone': row.get('Phone', '')
                        }
                    elif 'assistant' in role:
                        coach_data[key]['assistant_coach'] = {
                            'name': f"{row.get('First Name', '')} {row.get('Last Name', '')}",
                            'email': row.get('Email', ''),
                            'phone': row.get('Phone', '')
                        }
                        
            return coach_data
            
        except Exception as e:
            logger.error(f"Error loading coach data: {e}")
            return {}
            
    def process_game_card(self, 
                         game_card_path: str, 
                         enrollment_path: str = None,
                         volunteer_path: str = None) -> str:
        """
        Process game card by filling in upper division details
        
        Args:
            game_card_path: Path to the game card Excel file
            enrollment_path: Path to enrollment details (auto-finds if None)
            volunteer_path: Path to volunteer details (auto-finds if None)
            
        Returns:
            Path to the processed game card file
        """
        try:
            # Load data
            enrollment_df = self.load_enrollment_data(enrollment_path)
            self.coach_data = self.load_coach_data(volunteer_path)
            
            # Create backup if requested
            if self.backup_original:
                backup_path = game_card_path.replace('.xlsx', '_backup.xlsx')
                shutil.copy2(game_card_path, backup_path)
                logger.info(f"Created backup: {backup_path}")
            
            # Load the game card workbook
            wb = openpyxl.load_workbook(game_card_path)
            
            # Process each upper division
            for division in self.upper_divisions:
                logger.info(f"Processing division: {division}")
                self._process_division(wb, enrollment_df, division)
            
            # Save the processed file
            output_filename = os.path.basename(game_card_path).replace('.xlsx', f'{self.output_suffix}.xlsx')
            output_path = os.path.join(self.output_dir, output_filename)
            
            wb.save(output_path)
            logger.info(f"Saved processed game card to: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error processing game card: {e}")
            raise
            
    def _process_division(self, workbook: openpyxl.Workbook, 
                         enrollment_df: pd.DataFrame, 
                         division: str):
        """Process a single division in the game card"""
        
        # Filter for this division using the standard division column
        if 'Standard Division' in enrollment_df.columns:
            div_df = enrollment_df[enrollment_df['Standard Division'] == division].copy()
        else:
            # Fallback to the original method if Standard Division not available
            div_df = enrollment_df[enrollment_df['Division Name'].apply(
                lambda x: self._matches_target_division(str(x), division)
            )].copy()
        
        if div_df.empty:
            logger.warning(f"No players found for division {division}")
            return
            
        # Sort by team and jersey number
        div_df = div_df.sort_values(['Team Name', 'Player Jersey Number'], na_position='last')
        
        # Try to find the division sheet
        sheet_name = None
        for name in workbook.sheetnames:
            if division.lower() in name.lower():
                sheet_name = name
                break
                
        if not sheet_name:
            # Create new sheet for this division
            sheet_name = division
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
            logger.info(f"Created new sheet for {division}")
        else:
            sheet = workbook[sheet_name]
            
        # Clear existing data (keeping headers if they exist)
        self._clear_sheet_data(sheet)
        
        # Add division header
        self._add_division_header(sheet, division)
        
        # Process each team
        teams = div_df['Team Name'].unique()
        current_row = 4  # Start after header
        
        for team in sorted(teams):
            team_df = div_df[div_df['Team Name'] == team]
            current_row = self._add_team_section(sheet, team_df, team, division, current_row)
            current_row += 2  # Space between teams
            
    def _matches_target_division(self, division_name: str, target_division: str) -> bool:
        """
        Check if a division name matches a specific target division
        
        Args:
            division_name: The division name from the Excel file
            target_division: The target division (e.g., "16UG")
            
        Returns:
            True if this division matches the target
        """
        # Normalize the division name
        normalized = division_name.upper().replace(' ', '').replace('-', '')
        
        # Extract age and gender from target (e.g., "16UG" -> "16" and "G")
        age = target_division[:2]
        gender = target_division[-1]
        
        # Various possible formats
        patterns = [
            f"{age}U{gender}",           # 16UG
            f"{age}U{gender.lower()}",   # 16Ug
            f"U{age}{gender}",           # U16G
            f"U{age}{gender.lower()}",   # U16g
            f"{age}{gender}",            # 16G
        ]
        
        # Also check for full gender names
        if gender == 'G':
            patterns.extend([
                f"{age}UGIRLS", f"{age}UGIRL", f"U{age}GIRLS", f"U{age}GIRL",
                f"{age}GIRLS", f"{age}GIRL"
            ])
        elif gender == 'B':
            patterns.extend([
                f"{age}UBOYS", f"{age}UBOY", f"U{age}BOYS", f"U{age}BOY",
                f"{age}BOYS", f"{age}BOY"
            ])
        
        # Check if any pattern matches exactly (not just contains)
        for pattern in patterns:
            if pattern == normalized or normalized.startswith(pattern):
                return True
                
        return False
            
    def _clear_sheet_data(self, sheet):
        """Clear data from sheet while preserving structure"""
        # Find the last row with data
        max_row = sheet.max_row
        
        # Clear from row 4 onwards (assuming rows 1-3 are headers)
        for row in range(4, max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
                
    def _add_division_header(self, sheet, division: str):
        """Add formatted division header"""
        # Title
        sheet['A1'] = f"AREA 10V - LEAGUE GAME CARD"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:H1')
        
        # Division
        sheet['A2'] = f"Division: {division}"
        sheet['A2'].font = Font(bold=True, size=14)
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A2:H2')
        
        # Date
        sheet['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
        sheet['A3'].font = Font(italic=True, size=10)
        
    def _add_team_section(self, sheet, team_df: pd.DataFrame, 
                         team: str, division: str, start_row: int) -> int:
        """
        Add a team section to the sheet
        
        Returns:
            The next available row number
        """
        current_row = start_row
        
        # Team header
        sheet[f'A{current_row}'] = f"Team: {team}"
        sheet[f'A{current_row}'].font = Font(bold=True, size=12)
        sheet.merge_cells(f'A{current_row}:D{current_row}')
        
        # Add coach info
        coach_key = f"{division}_{team}"
        if coach_key in self.coach_data:
            current_row += 1
            coach_info = self.coach_data[coach_key]
            
            if 'head_coach' in coach_info:
                coach = coach_info['head_coach']
                sheet[f'A{current_row}'] = "Head Coach:"
                sheet[f'B{current_row}'] = coach['name']
                sheet[f'D{current_row}'] = coach.get('phone', '')
                sheet[f'F{current_row}'] = coach.get('email', '')
                current_row += 1
                
            if 'assistant_coach' in coach_info:
                coach = coach_info['assistant_coach']
                sheet[f'A{current_row}'] = "Asst Coach:"
                sheet[f'B{current_row}'] = coach['name']
                sheet[f'D{current_row}'] = coach.get('phone', '')
                sheet[f'F{current_row}'] = coach.get('email', '')
                current_row += 1
                
        current_row += 1
        
        # Player header
        headers = ['Jersey #', 'Player Name', 'Birth Date', 'Parent Name', 'Parent Email']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        current_row += 1
        
        # Add players
        for _, player in team_df.iterrows():
            sheet[f'A{current_row}'] = player.get('Player Jersey Number', '')
            sheet[f'B{current_row}'] = f"{player['Player First Name']} {player['Player Last Name']}"
            
            # Format birth date
            birth_date = player.get('Player Birth Date', '')
            if pd.notna(birth_date):
                try:
                    if isinstance(birth_date, str):
                        birth_date = pd.to_datetime(birth_date).strftime('%m/%d/%Y')
                    else:
                        birth_date = birth_date.strftime('%m/%d/%Y')
                except:
                    pass
            sheet[f'C{current_row}'] = birth_date
            
            sheet[f'D{current_row}'] = f"{player.get('Account First Name', '')} {player.get('Account Last Name', '')}"
            sheet[f'E{current_row}'] = player.get('User Email', '')
            
            current_row += 1
            
        # Add borders to the player section
        self._add_borders(sheet, start_row + 3, current_row - 1, 1, 5)
        
        return current_row
        
    def _add_borders(self, sheet, start_row: int, end_row: int, 
                    start_col: int, end_col: int):
        """Add borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row=row, column=col).border = thin_border
                
    def generate_summary_report(self, enrollment_df: pd.DataFrame = None) -> Dict:
        """Generate summary statistics for upper divisions"""
        if enrollment_df is None:
            enrollment_df = self.load_enrollment_data()
            
        summary = {
            'total_players': 0,
            'divisions': {},
            'division_name_mapping': self.division_mapping
        }
        
        # Use Standard Division if available, otherwise fall back to matching
        if 'Standard Division' in enrollment_df.columns:
            for division in self.upper_divisions:
                div_df = enrollment_df[enrollment_df['Standard Division'] == division]
                
                div_summary = {
                    'total_players': len(div_df),
                    'teams': div_df['Team Name'].nunique(),
                    'players_with_jerseys': div_df['Player Jersey Number'].notna().sum(),
                    'teams_list': div_df.groupby('Team Name').size().to_dict(),
                    'actual_division_names': div_df['Division Name'].unique().tolist()
                }
                
                summary['divisions'][division] = div_summary
                summary['total_players'] += len(div_df)
        else:
            # Fallback method
            for division in self.upper_divisions:
                div_df = enrollment_df[enrollment_df['Division Name'].apply(
                    lambda x: self._matches_target_division(str(x), division)
                )]
                
                div_summary = {
                    'total_players': len(div_df),
                    'teams': div_df['Team Name'].nunique(),
                    'players_with_jerseys': div_df['Player Jersey Number'].notna().sum(),
                    'teams_list': div_df.groupby('Team Name').size().to_dict(),
                    'actual_division_names': div_df['Division Name'].unique().tolist()
                }
                
                summary['divisions'][division] = div_summary
                summary['total_players'] += len(div_df)
            
        return summary