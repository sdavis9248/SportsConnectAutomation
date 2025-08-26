"""
Game Card Processor for AYSO League Game Cards
Fills upper division details from Enrollment Details report
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import shutil
import glob

logger = logging.getLogger(__name__)


class GameCardProcessor:
    """Process league game cards by filling upper division details from Enrollment Details"""
    
    UPPER_DIVISIONS = ['16UG', '16UB', '19UG', '19UB']
    
    # Column mappings for data extraction
    ENROLLMENT_COLUMNS = {
        'player_first': 'Player First Name',
        'player_last': 'Player Last Name',
        'player_id': 'Player ID',
        'division': 'Division Name',
        'team': 'Team Name',
        'team_number': 'Team Number',
        'jersey': 'Player Jersey Number',
        'birth_date': 'Player Birth Date',
        'order_no': 'Order No',
        'payment_status': 'Order Payment Status',
        'parent_first': 'Account First Name',
        'parent_last': 'Account Last Name',
        'parent_email': 'User Email',
        'registration_status': 'Registration Status'
    }
    
    def __init__(self, config):
        """
        Initialize Game Card Processor
        
        Args:
            config: Configuration manager instance
        """
        self.config = config
        self.download_dir = config.get('download_dir', 'data/downloads')
        self.output_dir = config.get('game_card_output_dir', 'data/game_cards')
        
        # Create output directory if it doesn't exist
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        game_card_config = config.get('game_card_config', {})
        self.upper_divisions = game_card_config.get('upper_divisions', self.UPPER_DIVISIONS)
        self.backup_original = game_card_config.get('backup_original', True)
        # Remove the output_suffix as we're not creating new files anymore
        # self.output_suffix = game_card_config.get('output_suffix', '_processed')
        
        # Coach data cache (would be populated from volunteer report)
        self.coach_data = {}
        
        # Division name mapping (actual name -> standard name)
        self.division_mapping = {}
        
    def find_latest_enrollment_file(self) -> Optional[str]:
        """Find the latest Enrollment Details file"""
        pattern = os.path.join(self.download_dir, "Enrollment_Details*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            logger.error("No Enrollment Details file found")
            return None
            
        # Get the most recent file
        latest_file = max(files, key=os.path.getmtime)
        logger.info(f"Found enrollment file: {latest_file}")
        return latest_file
        
    def find_latest_volunteer_file(self) -> Optional[str]:
        """Find the latest Volunteer Details file for coach information"""
        pattern = os.path.join(self.download_dir, "Volunteer_Details*.xlsx")
        files = glob.glob(pattern)
        
        if files:
            return max(files, key=os.path.getmtime)
        return None
        
    def load_enrollment_data(self, file_path: str = None) -> pd.DataFrame:
        """
        Load and prepare enrollment data
        
        Args:
            file_path: Path to enrollment file (auto-finds if None)
            
        Returns:
            Prepared DataFrame with enrollment data
        """
        if not file_path:
            file_path = self.find_latest_enrollment_file()
            
        if not file_path:
            raise FileNotFoundError("No enrollment file found")
            
        logger.info(f"Loading enrollment data from: {file_path}")
        df = pd.read_excel(file_path)
        
        # Standardize column names
        df.columns = df.columns.str.strip()
        
        # Debug: Show unique division names
        unique_divisions = df['Division Name'].unique()
        logger.info(f"Unique divisions in file: {unique_divisions}")
        
        # Build division mapping
        self._build_division_mapping(unique_divisions)
        
        # Filter for upper divisions - check if any of our target divisions are contained in the division name
        df_upper = df[df['Division Name'].apply(lambda x: self._is_upper_division(str(x)))].copy()
        
        # Add standardized division column for easier processing
        df_upper['Standard Division'] = df_upper['Division Name'].apply(
            lambda x: self.division_mapping.get(str(x), str(x))
        )
        
        # Filter out players who are not on a team (Unallocated)
        if 'Team Name' in df_upper.columns:
            df_upper = df_upper[df_upper['Team Name'] != 'Unallocated']
            df_upper = df_upper[df_upper['Team Name'].notna()]  # Also remove null team names
        
        logger.info(f"Found {len(df_upper)} players in upper divisions with team assignments")
        logger.info(f"Division breakdown: {df_upper['Standard Division'].value_counts().to_dict()}")
        
        return df_upper
        
    def _build_division_mapping(self, division_names):
        """
        Build a mapping from actual division names to standard names
        
        Args:
            division_names: List of division names from the Excel file
        """
        self.division_mapping = {}
        
        for div_name in division_names:
            if pd.isna(div_name):
                continue
                
            div_str = str(div_name)
            for target_div in self.upper_divisions:
                if self._matches_target_division(div_str, target_div):
                    self.division_mapping[div_str] = target_div
                    logger.debug(f"Mapped '{div_str}' to '{target_div}'")
                    break
                    
    def _is_upper_division(self, division_name: str) -> bool:
        """
        Check if a division name represents an upper division
        
        Args:
            division_name: The division name from the Excel file
            
        Returns:
            True if this is an upper division
        """
        # Normalize the division name
        normalized = division_name.upper().replace(' ', '').replace('-', '')
        
        # Check each target division
        for target_div in self.upper_divisions:
            # Extract age and gender from target (e.g., "16UG" -> "16" and "G")
            age = target_div[:2]
            gender = target_div[-1]
            
            # Various possible formats
            patterns = [
                f"{age}U{gender}",           # 16UG
                f"{age}U{gender.lower()}",   # 16Ug
                f"U{age}{gender}",           # U16G
                f"U{age}{gender.lower()}",   # U16g
                f"{age}{gender}",            # 16G
                f"{age}U",                   # 16U (if gender specified elsewhere)
            ]
            
            # Also check for full gender names
            if gender == 'G':
                patterns.extend([
                    f"{age}UGIRLS", f"{age}UGIRL", f"U{age}GIRLS", f"U{age}GIRL",
                    f"{age}GIRLS", f"{age}GIRL"
                ])
            elif gender == 'B':
                patterns.extend([
                    f"{age}UBOYS", f"{age}UBOY", f"U{age}BOYS", f"U{age}BOY",
                    f"{age}BOYS", f"{age}BOY"
                ])
            
            # Check if any pattern matches
            for pattern in patterns:
                if pattern in normalized:
                    logger.debug(f"Division '{division_name}' matched as {target_div}")
                    return True
                    
        return False
        
    def load_coach_data(self, file_path: str = None) -> Dict[str, Dict]:
        """
        Load coach data from volunteer report
        
        Args:
            file_path: Path to volunteer file (auto-finds if None)
            
        Returns:
            Dictionary mapping team identifiers to coach info
        """
        if not file_path:
            file_path = self.find_latest_volunteer_file()
            
        if not file_path:
            logger.warning("No volunteer file found - coach data will be empty")
            return {}
            
        try:
            logger.info(f"Loading coach data from: {file_path}")
            df = pd.read_excel(file_path)
            
            # Filter for coaches and assistant coaches
            coach_df = df[df['Volunteer Role'].str.contains('Coach', case=False, na=False)]
            
            coach_data = {}
            for _, row in coach_df.iterrows():
                division = row.get('Division Name', '')
                team = row.get('Team Name', '')
                
                if division and team:
                    key = f"{division}_{team}"
                    role = row.get('Volunteer Role', '').lower()
                    
                    if key not in coach_data:
                        coach_data[key] = {}
                    
                    if 'head' in role or role == 'coach':
                        coach_data[key]['head_coach'] = {
                            'name': f"{row.get('First Name', '')} {row.get('Last Name', '')}",
                            'email': row.get('Email', ''),
                            'phone': row.get('Phone', '')
                        }
                    elif 'assistant' in role:
                        coach_data[key]['assistant_coach'] = {
                            'name': f"{row.get('First Name', '')} {row.get('Last Name', '')}",
                            'email': row.get('Email', ''),
                            'phone': row.get('Phone', '')
                        }
                        
            return coach_data
            
        except Exception as e:
            logger.error(f"Error loading coach data: {e}")
            return {}
            
    def process_game_card(self, 
                         game_card_path: str, 
                         enrollment_path: str = None,
                         volunteer_path: str = None) -> str:
        """
        Process game card by filling in upper division details
        Works directly on the template file to preserve graphics
        
        Args:
            game_card_path: Path to the game card Excel file
            enrollment_path: Path to enrollment details (auto-finds if None)
            volunteer_path: Path to volunteer details (auto-finds if None)
            
        Returns:
            Path to the processed game card file
        """
        try:
            # Load data
            enrollment_df = self.load_enrollment_data(enrollment_path)
            self.coach_data = self.load_coach_data(volunteer_path)
            
            # Create backup if requested
            if self.backup_original:
                backup_path = game_card_path.replace('.xlsx', '_backup.xlsx')
                shutil.copy2(game_card_path, backup_path)
                logger.info(f"Created backup: {backup_path}")
            
            # Load the game card workbook
            wb = openpyxl.load_workbook(game_card_path)
            logger.info(f"Loaded workbook from: {game_card_path}")
            
            # Process each upper division
            divisions_processed = 0
            for division in self.upper_divisions:
                logger.info(f"Processing division: {division}")
                if self._process_division(wb, enrollment_df, division):
                    divisions_processed += 1
            
            if divisions_processed == 0:
                logger.warning("No divisions were processed")
                return None
            
            # Save the workbook back to the original file
            wb.save(game_card_path)
            logger.info(f"Saved processed game card to: {game_card_path}")
            
            # Close the workbook
            wb.close()
            
            return game_card_path
            
        except Exception as e:
            logger.error(f"Error processing game card: {e}")
            raise
            
    def _process_division(self, workbook: openpyxl.Workbook, 
                         enrollment_df: pd.DataFrame, 
                         division: str) -> bool:
        """
        Process a single division in the game card
        
        Returns:
            True if division was processed successfully
        """
        
        # Filter for this division using the standard division column
        if 'Standard Division' in enrollment_df.columns:
            div_df = enrollment_df[enrollment_df['Standard Division'] == division].copy()
        else:
            # Fallback to the original method if Standard Division not available
            div_df = enrollment_df[enrollment_df['Division Name'].apply(
                lambda x: self._matches_target_division(str(x), division)
            )].copy()
        
        if div_df.empty:
            logger.warning(f"No players found for division {division}")
            return False
            
        # Sort by team and jersey number
        div_df = div_df.sort_values(['Team Name', 'Player Jersey Number'], na_position='last')
        
        # Check if the division sheet exists
        sheet_name = f"Game Card {division}"
        
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found. Please create it by copying the 'Game Card' template.")
            return False
            
        sheet = workbook[sheet_name]
        logger.info(f"Using sheet: {sheet_name}")
        
        # Check if overflow sheet exists
        overflow_sheet_name = f"{sheet_name}-2"
        has_overflow_sheet = overflow_sheet_name in workbook.sheetnames
        overflow_sheet = workbook[overflow_sheet_name] if has_overflow_sheet else None
        
        if has_overflow_sheet:
            logger.info(f"Found overflow sheet: {overflow_sheet_name}")
            # Clear all sections on overflow sheet
            for position in range(1, 4):
                self._clear_game_card_section(overflow_sheet, position)
        
        # Process teams in the division
        teams = div_df['Team Name'].unique()
        
        # Clear all 3 sections on main sheet first
        for position in range(1, 4):
            self._clear_game_card_section(sheet, position)
        
        # Process teams
        teams_processed = 0
        current_sheet = sheet
        current_position = 1
        
        for team_idx, team_name in enumerate(teams):
            team_df = div_df[div_df['Team Name'] == team_name]
            
            # Check if we need to move to overflow sheet
            if current_position > 3:
                if has_overflow_sheet:
                    current_sheet = overflow_sheet
                    current_position = 1
                    logger.info(f"Moving to overflow sheet for remaining teams")
                else:
                    logger.warning(f"No more space for team {team_name}. Need overflow sheet '{overflow_sheet_name}'")
                    break
            
            # Process this team with potential player overflow
            players_remaining = self._fill_game_card_team_with_overflow(
                current_sheet, 
                overflow_sheet if has_overflow_sheet else None,
                team_df, 
                division, 
                team_name, 
                current_position,
                team_idx
            )
            
            if players_remaining > 0 and not has_overflow_sheet:
                logger.warning(f"Team {team_name} has {players_remaining} players that don't fit. Need overflow sheet.")
            
            teams_processed += 1
            current_position += 1
            
        if teams_processed < len(teams):
            logger.warning(f"Division {division} has {len(teams)} teams, but only {teams_processed} were processed")
            
        return teams_processed > 0
    
    def _fill_game_card_team_with_overflow(self, main_sheet, overflow_sheet, team_df: pd.DataFrame, 
                                          division: str, team_name: str, position: int, team_idx: int) -> int:
        """
        Fill in team information with overflow handling
        
        Args:
            main_sheet: The main worksheet to fill
            overflow_sheet: The overflow worksheet (can be None)
            team_df: DataFrame with team players
            division: Division name
            team_name: Team name
            position: Position on main sheet (1, 2, or 3)
            team_idx: Index of this team (0-based)
            
        Returns:
            Number of players that didn't fit
        """
        # First, fill the main sheet
        self._clear_game_card_section(main_sheet, position)
        
        # Define the cell locations
        team_sections = {
            1: {
                'division': 'F4',
                'team_name': 'F5',
                'coach': 'D7',
                'assistant': 'G7',
                'team_id': 'J4',
                'player_start_row': 11,
                'jersey_col': 'B',
                'name_col': 'C'
            },
            2: {
                'division': 'S4',
                'team_name': 'S5',
                'coach': 'Q7',
                'assistant': 'T7',
                'team_id': 'W4',
                'player_start_row': 11,
                'jersey_col': 'O',
                'name_col': 'P'
            },
            3: {
                'division': 'AF4',
                'team_name': 'AF5',
                'coach': 'AD7',
                'assistant': 'AG7',
                'team_id': 'AJ4',
                'player_start_row': 11,
                'jersey_col': 'AB',
                'name_col': 'AC'
            }
        }
        
        section = team_sections[position]
        
        # Fill in team header information on main sheet
        main_sheet[section['division']] = division
        main_sheet[section['team_name']] = team_name
        
        # Get coach information if available
        coach_key = f"{division}_{team_name}"
        if coach_key in self.coach_data:
            coach_info = self.coach_data[coach_key]
            if 'head_coach' in coach_info:
                main_sheet[section['coach']] = coach_info['head_coach']['name']
            if 'assistant_coach' in coach_info:
                main_sheet[section['assistant']] = coach_info['assistant_coach']['name']
        
        # Fill in players
        max_players_per_section = 20
        players_filled = 0
        current_row = section['player_start_row']
        
        for idx, (_, player) in enumerate(team_df.iterrows()):
            if players_filled < max_players_per_section:
                # Fill on main sheet
                jersey_num = player.get('Player Jersey Number', '')
                if pd.notna(jersey_num):
                    main_sheet[f"{section['jersey_col']}{current_row}"] = jersey_num
                    
                player_name = f"{player['Player First Name']} {player['Player Last Name']}"
                main_sheet[f"{section['name_col']}{current_row}"] = player_name
                
                current_row += 1
                players_filled += 1
            else:
                # Need overflow
                if overflow_sheet is not None:
                    # Calculate overflow position
                    # Use same position if team index < 3, otherwise next available
                    overflow_position = position if team_idx < 3 else ((team_idx - 3) % 3) + 1
                    
                    # Only fill header on first overflow player
                    if players_filled == max_players_per_section:
                        logger.info(f"Team {team_name} has more than {max_players_per_section} players, using overflow sheet")
                        overflow_section = team_sections[overflow_position]
                        
                        # Fill team info on overflow sheet
                        overflow_sheet[overflow_section['division']] = division
                        overflow_sheet[overflow_section['team_name']] = f"{team_name} (cont.)"
                        
                        if coach_key in self.coach_data:
                            coach_info = self.coach_data[coach_key]
                            if 'head_coach' in coach_info:
                                overflow_sheet[overflow_section['coach']] = coach_info['head_coach']['name']
                            if 'assistant_coach' in coach_info:
                                overflow_sheet[overflow_section['assistant']] = coach_info['assistant_coach']['name']
                    
                    # Fill player on overflow sheet
                    overflow_row = overflow_section['player_start_row'] + (players_filled - max_players_per_section)
                    
                    jersey_num = player.get('Player Jersey Number', '')
                    if pd.notna(jersey_num):
                        overflow_sheet[f"{overflow_section['jersey_col']}{overflow_row}"] = jersey_num
                        
                    player_name = f"{player['Player First Name']} {player['Player Last Name']}"
                    overflow_sheet[f"{overflow_section['name_col']}{overflow_row}"] = player_name
                    
                    players_filled += 1
                    
                    # Check if overflow is also full
                    if (players_filled - max_players_per_section) >= max_players_per_section:
                        logger.warning(f"Team {team_name} has more than {max_players_per_section * 2} players!")
                        break
                else:
                    # No overflow sheet available
                    break
        
        players_remaining = len(team_df) - players_filled
        
        if players_filled <= max_players_per_section:
            logger.info(f"Filled team {position}: {team_name} with {players_filled} players")
        else:
            logger.info(f"Filled team {position}: {team_name} with {max_players_per_section} players on main sheet and {players_filled - max_players_per_section} on overflow")
            
        return players_remaining
            
    def _matches_target_division(self, division_name: str, target_division: str) -> bool:
        """
        Check if a division name matches a specific target division
        
        Args:
            division_name: The division name from the Excel file
            target_division: The target division (e.g., "16UG")
            
        Returns:
            True if this division matches the target
        """
        # Normalize the division name
        normalized = division_name.upper().replace(' ', '').replace('-', '')
        
        # Extract age and gender from target (e.g., "16UG" -> "16" and "G")
        age = target_division[:2]
        gender = target_division[-1]
        
        # Various possible formats
        patterns = [
            f"{age}U{gender}",           # 16UG
            f"{age}U{gender.lower()}",   # 16Ug
            f"U{age}{gender}",           # U16G
            f"U{age}{gender.lower()}",   # U16g
            f"{age}{gender}",            # 16G
        ]
        
        # Also check for full gender names
        if gender == 'G':
            patterns.extend([
                f"{age}UGIRLS", f"{age}UGIRL", f"U{age}GIRLS", f"U{age}GIRL",
                f"{age}GIRLS", f"{age}GIRL"
            ])
        elif gender == 'B':
            patterns.extend([
                f"{age}UBOYS", f"{age}UBOY", f"U{age}BOYS", f"U{age}BOY",
                f"{age}BOYS", f"{age}BOY"
            ])
        
        # Check if any pattern matches exactly (not just contains)
        for pattern in patterns:
            if pattern == normalized or normalized.startswith(pattern):
                return True
                
        return False
            
    def _clear_sheet_data(self, sheet):
        """Clear data from sheet while preserving structure"""
        # Find the last row with data
        max_row = sheet.max_row
        
        # Clear from row 4 onwards (assuming rows 1-3 are headers)
        for row in range(4, max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
                
    def _add_division_header(self, sheet, division: str):
        """Add formatted division header"""
        # Title
        sheet['A1'] = f"AREA 10V - LEAGUE GAME CARD"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:H1')
        
        # Division
        sheet['A2'] = f"Division: {division}"
        sheet['A2'].font = Font(bold=True, size=14)
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A2:H2')
        
        # Date
        sheet['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
        sheet['A3'].font = Font(italic=True, size=10)
        
    def _fill_game_card_team_with_overflow(self, main_sheet, overflow_sheet, team_df: pd.DataFrame, 
                                          division: str, team_name: str, position: int, team_idx: int) -> int:
        """
        Fill in team information with overflow handling
        
        Args:
            main_sheet: The main worksheet to fill
            overflow_sheet: The overflow worksheet (can be None)
            team_df: DataFrame with team players
            division: Division name
            team_name: Team name
            position: Position on main sheet (1, 2, or 3)
            team_idx: Index of this team (0-based)
            
        Returns:
            Number of players that didn't fit
        """
        # First, fill the main sheet
        self._clear_game_card_section(main_sheet, position)
        
        # Define the cell locations
        team_sections = {
            1: {
                'division': 'F4',
                'team_name': 'F5',
                'coach': 'D7',
                'assistant': 'G7',
                'team_id': 'J4',
                'player_start_row': 11,
                'jersey_col': 'B',
                'name_col': 'C'
            },
            2: {
                'division': 'S4',
                'team_name': 'S5',
                'coach': 'Q7',
                'assistant': 'T7',
                'team_id': 'W4',
                'player_start_row': 11,
                'jersey_col': 'O',
                'name_col': 'P'
            },
            3: {
                'division': 'AF4',
                'team_name': 'AF5',
                'coach': 'AD7',
                'assistant': 'AG7',
                'team_id': 'AJ4',
                'player_start_row': 11,
                'jersey_col': 'AB',
                'name_col': 'AC'
            }
        }
        
        section = team_sections[position]
        
        # Fill in team header information on main sheet
        main_sheet[section['division']] = division
        main_sheet[section['team_name']] = team_name
        
        # Get coach information if available
        coach_key = f"{division}_{team_name}"
        if coach_key in self.coach_data:
            coach_info = self.coach_data[coach_key]
            if 'head_coach' in coach_info:
                main_sheet[section['coach']] = coach_info['head_coach']['name']
            if 'assistant_coach' in coach_info:
                main_sheet[section['assistant']] = coach_info['assistant_coach']['name']
        
        # Fill in players
        max_players_per_section = 20
        players_filled = 0
        current_row = section['player_start_row']
        
        for idx, (_, player) in enumerate(team_df.iterrows()):
            if players_filled < max_players_per_section:
                # Fill on main sheet
                jersey_num = player.get('Player Jersey Number', '')
                if pd.notna(jersey_num):
                    main_sheet[f"{section['jersey_col']}{current_row}"] = jersey_num
                    
                player_name = f"{player['Player First Name']} {player['Player Last Name']}"
                main_sheet[f"{section['name_col']}{current_row}"] = player_name
                
                current_row += 1
                players_filled += 1
            else:
                # Need overflow
                if overflow_sheet is not None:
                    # Calculate overflow position
                    # Use same position if team index < 3, otherwise next available
                    overflow_position = position if team_idx < 3 else ((team_idx - 3) % 3) + 1
                    
                    # Only fill header on first overflow player
                    if players_filled == max_players_per_section:
                        logger.info(f"Team {team_name} has more than {max_players_per_section} players, using overflow sheet")
                        overflow_section = team_sections[overflow_position]
                        
                        # Fill team info on overflow sheet
                        overflow_sheet[overflow_section['division']] = division
                        overflow_sheet[overflow_section['team_name']] = f"{team_name} (cont.)"
                        
                        if coach_key in self.coach_data:
                            coach_info = self.coach_data[coach_key]
                            if 'head_coach' in coach_info:
                                overflow_sheet[overflow_section['coach']] = coach_info['head_coach']['name']
                            if 'assistant_coach' in coach_info:
                                overflow_sheet[overflow_section['assistant']] = coach_info['assistant_coach']['name']
                    
                    # Fill player on overflow sheet
                    overflow_row = overflow_section['player_start_row'] + (players_filled - max_players_per_section)
                    
                    jersey_num = player.get('Player Jersey Number', '')
                    if pd.notna(jersey_num):
                        overflow_sheet[f"{overflow_section['jersey_col']}{overflow_row}"] = jersey_num
                        
                    player_name = f"{player['Player First Name']} {player['Player Last Name']}"
                    overflow_sheet[f"{overflow_section['name_col']}{overflow_row}"] = player_name
                    
                    players_filled += 1
                    
                    # Check if overflow is also full
                    if (players_filled - max_players_per_section) >= max_players_per_section:
                        logger.warning(f"Team {team_name} has more than {max_players_per_section * 2} players!")
                        break
                else:
                    # No overflow sheet available
                    break
        
        players_remaining = len(team_df) - players_filled
        
        if players_filled <= max_players_per_section:
            logger.info(f"Filled team {position}: {team_name} with {players_filled} players")
        else:
            logger.info(f"Filled team {position}: {team_name} with {max_players_per_section} players on main sheet and {players_filled - max_players_per_section} on overflow")
            
        return players_remaining

    def _clear_game_card_section(self, sheet, position: int):
        """
        Clear data from a specific team section on the game card
        
        Args:
            sheet: The worksheet to clear
            position: Position to clear (1, 2, or 3)
        """
        # Define the cell locations for each of the 3 team sections
        team_sections = {
            1: {
                'division': 'F4',
                'team_name': 'F5',
                'coach': 'D7',
                'assistant': 'G7',
                'team_id': 'J4',
                'player_start_row': 11,
                'player_end_row': 30,  # 20 players max
                'jersey_col': 'B',
                'name_col': 'C'
            },
            2: {
                'division': 'S4',
                'team_name': 'S5',
                'coach': 'Q7',
                'assistant': 'T7',
                'team_id': 'W4',
                'player_start_row': 11,
                'player_end_row': 30,
                'jersey_col': 'O',
                'name_col': 'P'
            },
            3: {
                'division': 'AF4',
                'team_name': 'AF5',
                'coach': 'AD7',
                'assistant': 'AG7',
                'team_id': 'AJ4',
                'player_start_row': 11,
                'player_end_row': 30,
                'jersey_col': 'AB',
                'name_col': 'AC'
            }
        }
        
        if position not in team_sections:
            return
            
        section = team_sections[position]
        
        # Clear team header information
        sheet[section['division']] = ''
        sheet[section['team_name']] = ''
        sheet[section['coach']] = ''
        sheet[section['assistant']] = ''
        sheet[section['team_id']] = ''
        
        # Clear player roster
        for row in range(section['player_start_row'], section['player_end_row'] + 1):
            sheet[f"{section['jersey_col']}{row}"] = ''
            sheet[f"{section['name_col']}{row}"] = ''
        """
        Clear data from a specific team section on the game card
        
        Args:
            sheet: The worksheet to clear
            position: Position to clear (1, 2, or 3)
        """
        # Define the cell locations for each of the 3 team sections
        team_sections = {
            1: {
                'division': 'F4',
                'team_name': 'F5',
                'coach': 'D7',
                'assistant': 'G7',
                'team_id': 'J4',
                'player_start_row': 11,
                'player_end_row': 30,  # 20 players max
                'jersey_col': 'B',
                'name_col': 'C'
            },
            2: {
                'division': 'S4',
                'team_name': 'S5',
                'coach': 'Q7',
                'assistant': 'T7',
                'team_id': 'W4',
                'player_start_row': 11,
                'player_end_row': 30,
                'jersey_col': 'O',
                'name_col': 'P'
            },
            3: {
                'division': 'AF4',
                'team_name': 'AF5',
                'coach': 'AD7',
                'assistant': 'AG7',
                'team_id': 'AJ4',
                'player_start_row': 11,
                'player_end_row': 30,
                'jersey_col': 'AB',
                'name_col': 'AC'
            }
        }
        
        if position not in team_sections:
            return
            
        section = team_sections[position]
        
        # Clear team header information
        sheet[section['division']] = ''
        sheet[section['team_name']] = ''
        sheet[section['coach']] = ''
        sheet[section['assistant']] = ''
        sheet[section['team_id']] = ''
        
        # Clear player roster
        for row in range(section['player_start_row'], section['player_end_row'] + 1):
            sheet[f"{section['jersey_col']}{row}"] = ''
            sheet[f"{section['name_col']}{row}"] = ''
    
    def _fill_game_card_team(self, sheet, team_df: pd.DataFrame, division: str, team_name: str, position: int):
        """
        Fill in team information on the game card template
        
        Args:
            sheet: The worksheet to fill
            team_df: DataFrame with team players
            division: Division name
            team_name: Team name
            position: Position on sheet (1, 2, or 3)
        """
        # First clear the section
        self._clear_game_card_section(sheet, position)
        
        # Define the cell locations for each of the 3 team sections
        team_sections = {
            1: {
                'division': 'F4',
                'team_name': 'F5',
                'coach': 'D7',
                'assistant': 'G7',
                'team_id': 'J4',
                'player_start_row': 11,
                'jersey_col': 'B',
                'name_col': 'C'
            },
            2: {
                'division': 'S4',
                'team_name': 'S5',
                'coach': 'Q7',
                'assistant': 'T7',
                'team_id': 'W4',
                'player_start_row': 11,
                'jersey_col': 'O',
                'name_col': 'P'
            },
            3: {
                'division': 'AF4',
                'team_name': 'AF5',
                'coach': 'AD7',
                'assistant': 'AG7',
                'team_id': 'AJ4',
                'player_start_row': 11,
                'jersey_col': 'AB',
                'name_col': 'AC'
            }
        }
        
        if position not in team_sections:
            logger.warning(f"Invalid position {position} for team section")
            return
            
        section = team_sections[position]
        
        # Fill in team header information
        sheet[section['division']] = division
        sheet[section['team_name']] = team_name
        
        # Get coach information if available
        coach_key = f"{division}_{team_name}"
        if coach_key in self.coach_data:
            coach_info = self.coach_data[coach_key]
            if 'head_coach' in coach_info:
                sheet[section['coach']] = coach_info['head_coach']['name']
            if 'assistant_coach' in coach_info:
                sheet[section['assistant']] = coach_info['assistant_coach']['name']
        
        # Fill in player roster
        current_row = section['player_start_row']
        for _, player in team_df.iterrows():
            # Jersey number
            jersey_num = player.get('Player Jersey Number', '')
            if pd.notna(jersey_num):
                sheet[f"{section['jersey_col']}{current_row}"] = jersey_num
                
            # Player name
            player_name = f"{player['Player First Name']} {player['Player Last Name']}"
            sheet[f"{section['name_col']}{current_row}"] = player_name
            
            current_row += 1
            
            # Stop if we've filled too many rows (template might have a limit)
            if current_row > section['player_start_row'] + 19:  # 20 players max (rows 11-30)
                logger.warning(f"Too many players for team {team_name}, some may be cut off")
                break
        
    def _add_borders(self, sheet, start_row: int, end_row: int, 
                    start_col: int, end_col: int):
        """Add borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row=row, column=col).border = thin_border
                
    def generate_summary_report(self, enrollment_df: pd.DataFrame = None) -> Dict:
        """Generate summary statistics for upper divisions"""
        if enrollment_df is None:
            enrollment_df = self.load_enrollment_data()
            
        summary = {
            'total_players': 0,
            'divisions': {},
            'division_name_mapping': self.division_mapping
        }
        
        # Use Standard Division if available, otherwise fall back to matching
        if 'Standard Division' in enrollment_df.columns:
            for division in self.upper_divisions:
                div_df = enrollment_df[enrollment_df['Standard Division'] == division]
                
                div_summary = {
                    'total_players': len(div_df),
                    'teams': div_df['Team Name'].nunique(),
                    'players_with_jerseys': div_df['Player Jersey Number'].notna().sum(),
                    'teams_list': div_df.groupby('Team Name').size().to_dict(),
                    'actual_division_names': div_df['Division Name'].unique().tolist()
                }
                
                summary['divisions'][division] = div_summary
                summary['total_players'] += len(div_df)
        else:
            # Fallback method
            for division in self.upper_divisions:
                div_df = enrollment_df[enrollment_df['Division Name'].apply(
                    lambda x: self._matches_target_division(str(x), division)
                )]
                
                div_summary = {
                    'total_players': len(div_df),
                    'teams': div_df['Team Name'].nunique(),
                    'players_with_jerseys': div_df['Player Jersey Number'].notna().sum(),
                    'teams_list': div_df.groupby('Team Name').size().to_dict(),
                    'actual_division_names': div_df['Division Name'].unique().tolist()
                }
                
                summary['divisions'][division] = div_summary
                summary['total_players'] += len(div_df)
            
        return summary
    
    def create_game_card_instructions(self) -> str:
        """
        Create instructions for manual sheet preparation when graphics are involved
        
        Returns:
            Instructions text
        """
        instructions = """
GAME CARD PREPARATION INSTRUCTIONS
==================================

Due to graphics in the template, sheets need to be manually prepared:

1. Open the game card template file
2. For each upper division (16UG, 16UB, 19UG, 19UB):
   - Right-click on the "Game Card" sheet tab
   - Select "Move or Copy..."
   - Check "Create a copy"
   - Click OK
   - Rename the new sheet to "Game Card [DIVISION]" (e.g., "Game Card 16UG")

3. Save the file with all division sheets created

4. Run the processor with:
   python main.py --process-game-card

The processor will fill in the data on each prepared sheet while preserving graphics.

Alternative: Process one division at a time:
   python main.py --process-game-card-division 16UG
"""
        return instructions